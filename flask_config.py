import os, platform
from flask import Flask, render_template, redirect, request, url_for
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

system_name = str(platform.system())
if system_name == "Windows":
    slash = "\\"
    log_folder = "Log\\Test Log\\"
    attachment_folder = "Attachment"
    testplan_id_file = attachment_folder + "\\TestPlan_Database.xlsx"
else:
    slash = "/"
    log_folder = "Log/Test Log/"
    attachment_folder = "Attachment"
    testplan_id_file = attachment_folder + "/TestPlan_Database.xlsx"

def DefineTestPlan():
    testplan = False

    if request.method == "POST":
        testplan_name = request.form["name"]
        for file in os.listdir(log_folder):
            if file.startswith(testplan_name):
                testplan = True
    
    if bool(testplan) == True:
        wb = load_workbook(testplan_id_file)
        ws = wb.active

        last_row = ws.max_row
        for row in range(1,last_row):
            row+=1
            
            plan_name = ws.cell(row=row, column=1).value
            if plan_name == testplan_name:
                testplan_id = ws.cell(row=row, column=2).value
                break
    
    return (testplan, testplan_id)



@app.route('/home', methods=['GET', 'POST'])
def homepage():

    return render_template('index.html')

@app.route('/list', methods=['POST'])
def section_list():
    global testplan
    testplan = DefineTestPlan()

    if bool(testplan) == False:
        return "<div>test plan not found</div>"
    else:
        global testplan_id
        testplan_id = testplan[1]
        return redirect(url_for('show_list', testplan_id=testplan_id))

@app.route('/list/<testplan_id>')
def show_list(testplan_id):
    files = []
    list_data = {}

    for file in os.listdir(log_folder):
        if file.endswith(".xlsx"):
            if "/" in file:
                file_name = str(os.path.join(log_folder, file)).split(".xlsx")[0].split("%sLog%sTest Log%s" % (slash_sign, slash_sign, slash_sign))[1]
            elif "\\" in file:
                file_name = str(os.path.join(log_folder, file)).split(".xlsx")[0].split("\\Log\\Test Log\\")[1]
            else:
                slash_sign = " \ ".strip()
                if slash_sign in file:
                    file_name = str(os.path.join(log_folder, file)).split(".xlsx")[0].split("\Log\Test Log" + slash_sign)[1]
                else:
                    file_name = str(os.path.join(log_folder, file)).split(".xlsx")[0]

            files.append(file_name)

    files_len = len(files)
    for file in range(0,files_len):
        file+=1

        global filename
        filename = files[-file]
        print("filename" + str(filename))
        test_id = str(filename).split("_result_")[1]
        
        no = file
        
        test_date = "/".join([test_id[0:2], test_id[2:4], test_id[4:6]])
        test_time = ":".join([test_id[6:8], test_id[8:10], test_id[10:12]])
        date_time = "%s %s" % (test_date, test_time)
        
        passed_list = []
        failed_list = []
        
        log_repos = "%s.xlsx" % filename
        tc_log = load_workbook(log_repos)
        tc_sheet = tc_log.active

        tc_last_row = tc_sheet.max_row
        for row in range(1,tc_last_row):
            row+=1

            status = tc_sheet.cell(row=row, column=5).value
            if status == "Pass":
                passed_list.append(status)
            elif status == "Fail":
                failed_list.append(status)
        
        tc_pass = len(passed_list)
        tc_fail = len(failed_list)

        list_data[file] = {
            "no": no,
            "id": test_id,
            "name": date_time,
            "pass": tc_pass,
            "fail": tc_fail
        }

    return render_template('test_list.html', list_data=list_data, testplan_id=testplan_id)

@app.route('/details/<testplan_id>/<testcase_id>')
def section_details(testplan_id, testcase_id):
    if bool(testplan) == False:
        return "<div>test plan not found</div>"
    else:
        testcase_dict = {}
        passed_list = []
        failed_list = []
        current_section_output = {}
        
        log_repos = "%s.xlsx" % filename

        tc_log = load_workbook(log_repos)
        tc_sheet = tc_log.active

        tc_last_row = tc_sheet.max_row
        for row in range(1,tc_last_row):
            row+=1

            no = row-1
            menu = tc_sheet.cell(row=row, column=2).value
            division = tc_sheet.cell(row=row, column=3).value
            testcase = tc_sheet.cell(row=row, column=4).value
            status = tc_sheet.cell(row=row, column=5).value
            date = tc_sheet.cell(row=row, column=6).value

            if status == "Pass":
                passed_list.append(status)
            elif status == "Fail":
                failed_list.append(status)

            if bool(menu) == True:
                testcase_dict[no] = {
                    "no": no,
                    "menu": menu,
                    "division": division,
                    "testcase": testcase,
                    "status": status,
                    "date": date
                }

        project_name = tc_sheet.cell(row=2, column=10).value
        domain_name = tc_sheet.cell(row=2, column=11).value
        account = str(tc_sheet.cell(row=2, column=12).value) + "/" + str(tc_sheet.cell(row=2, column=13).value)
        version = tc_sheet.cell(row=2, column=14).value

        tc_log.save(log_repos)

        passed_tc = str(len(passed_list))
        failed_tc = str(len(failed_list))

        current_section_output = {
            "testcase_dict": testcase_dict,
            "passed_tc": passed_tc,
            "failed_tc": failed_tc,
            "project_name": project_name,
            "domain_name": domain_name,
            "account": account,
            "version": version
        }

        return render_template('test_details.html', current_section_output=current_section_output)

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0")