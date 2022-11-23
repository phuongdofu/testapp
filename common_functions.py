import time, sys, unittest, random, json, requests, openpyxl, testlink, platform, os, shutil
from unicodedata import name
from datetime import datetime
from selenium import webdriver
from random import randint
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoAlertPresentException, TimeoutException, WebDriverException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.alert import Alert
from openpyxl import Workbook, load_workbook
import testlink

chrome_options = Options()
chrome_options.add_argument("--start-maximized")

system_name = str(platform.system())
if system_name == "Windows":
    slash = "\\"
    json_file = os.path.dirname(os.path.realpath(__file__)) + "\\config.json"
    chromedriver_file = "chromedriver.exe"
    system_path = "C:\\Users\\Hanbiro\\"  
else:
    slash = "/"
    json_file = os.path.dirname(os.path.realpath(__file__)) + "/config.json"
    chromedriver_file = "chromedriver"
    system_path = ""

date_time = datetime.now().strftime("%Y/%m/%d, %H:%M:%S")
# section_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]

with open(json_file) as json_data_file:
    data = json.load(json_data_file)
    approval_dict = dict(data["approval"])
    archive_dict = dict(data["archive"])
    asset_dict = dict(data["asset"])
    board_dict = dict(data["board"])
    calendar_dict = dict(data["calendar_folder"])
    circular_dict = dict(data["circular"])
    clouddisk_dict = dict(data["clouddisk"])
    contact_dict = dict(data["contact"])
    expense_dict = dict(data["expense"])
    mail_dict = dict(data["mail"])
    project_dict = dict(data["project"])
    resource_dict =dict(data["resource"])
    diary_dict = dict(data["work_diary"])
    report_dict = dict(data["task_report"])
    todo_dict = dict(data["todo"])
    whisper_dict = dict(data["whisper"])
    approval_tc = dict(data["testcase_result"]["approval"])
    archive_tc = dict(data["testcase_result"]["archive"])
    asset_tc = dict(data["testcase_result"]["asset"])
    board_tc = dict(data["testcase_result"]["board"])
    calendar_tc = dict(data["testcase_result"]["calendar_folder"])
    circular_tc = dict(data["testcase_result"]["circular"])
    clouddisk_tc = dict(data["testcase_result"]["clouddisk"])
    contact_tc = dict(data["testcase_result"]["contact"])
    expense_tc = dict(data["testcase_result"]["expense"])
    mail_tc = dict(data["testcase_result"]["mail"])
    project_tc = dict(data["testcase_result"]["project"])
    resource_tc = dict(data["testcase_result"]["resource"])
    diary_tc = dict(data["testcase_result"]["work_diary"])
    report_tc = dict(data["testcase_result"]["task_report"])
    todo_tc = dict(data["testcase_result"]["todo"])
    whisper_tc = dict(data["testcase_result"]["whisper"])

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
       
class Files():
    local_path = os.path.dirname(os.path.realpath(__file__)) + slash

    execution_folder = ""
    log_folder = "%sLog%s" % (local_path, slash)
    test_log_folder = log_folder + "Test Log" + slash
    attachment_folder = "%sAttachment%s" % (local_path, slash)
    clouddisk_folder = "%sAttachment%sCloudDisk%s" % (local_path, slash, slash)
    download_folder = "%sDownloads" % system_path
    json_file = local_path + "config.json"

    chromedriver_path = local_path + chromedriver_file
    execution_log = log_folder + "execution_log.txt"
    testplan = log_folder + "testplan.txt"
    fail_log = execution_log.replace("execution_log", "fail_log")
    error_log = execution_log.replace("execution_log", "error_log")
    testcase_log = log_folder + "testcase_log_result.xlsx"
    collect_menu_log = log_folder + "select-menu-file.xlsx"
    section_history = log_folder + "section-history.xlsx"

    testplan_name = open(testplan, "r").read()
    #testcase_filename = "%s_result_%s.xlsx" % (testplan_name, section_id)
    #testcase_file  = test_log_folder + testcase_filename

    image_attachment = attachment_folder + "download.jpg"
    asset_import = attachment_folder + "Asset-SeleniumPython.xls"
    calendar_import = attachment_folder + "Calendar-SeleniumPython.xls"
    contact_import = attachment_folder + "Contact-SeleniumPython.xls"
    expense_import = attachment_folder + "Expense-SeleniumPython.xls"
    testplan_file = attachment_folder + "TestPlan_Database.xlsx"

    def ConfigFiles():
        if system_name == "Windows" and Files.myfolder == False:
            config_files = ["chromedriver.exe", 
                            "geckodriver.exe",
                            "chromedriver_talk.exe",
                            "main_functions.py", 
                            "run_files.py", 
                            "testapp_ui.py", 
                            "windows_messenger.py",
                            "config.json", 
                            "MN_groupware_auto.json"]

            for config_file in config_files:
                source_path = Files.system_path + "TestApp_v4" + slash + config_file
                destination_path = Files.local_path + slash + config_file
                shutil.copy(source_path, destination_path)

            folder_source_path = Files.system_path + slash + "TestApp_v4" + slash + "Attachment"
            folder_destination_path = Files.local_path + slash + "Attachment"
            shutil.copytree(folder_source_path, folder_destination_path)

class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    today = now.strftime("%Y/%m/%d")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]

    testcase_fail = "%sTest case status: fail%s" %  ('\033[91m', '\033[0m')
    testcase_pass = "%sTest case status: pass%s" %  ('\033[92m', '\033[0m')
    testcase_block = "%sTest case status: block%s" %  ('\033[93m', '\033[0m')
    
    hanbiro_title = "generated by selenium python at %s" % date_time
    title_edit = "Title is updated at %s" % date_time
    content_edit = "Content is updated at %s" % date_time
    hanbiro_content = "Content is written at %s" % date_time

    slash_list = [" \ ", " / "]

class Driver():
    def StartWebdriver():
        args = ["hide_console"]
        driver = webdriver.Chrome(Files.chromedriver_path, service_args=args)

        try:
            driver.maximize_window()
        except:
            Files.ConfigFiles()
            driver = webdriver.Chrome(Files.chromedriver_path, service_args=args)
            driver.maximize_window()
        
        return driver
        
    def StartTalkDriver():
        # Start the web driver
        service = webdriver.chrome.service.Service(Files.local_path + "%schromedriver_talk.exe" % slash)
        service.start()

        # start the app
        driver = webdriver.remote.webdriver.WebDriver(
            command_executor=service.service_url,
            desired_capabilities={
                'browserName': 'chrome',
                'goog:chromeOptions': {
                    'args': ['develop_mode'],
                    'binary': '%s\\AppData\\Local\\Programs\\hanbiro-talk\\HanbiroTalk2.exe' % Files.folder_execution,
                    'extensions': [],
                    'windowTypes': ['webview']},
                'platform': 'ANY',
                'version': ''},
            browser_profile=None,
            proxy=None,
            keep_alive=False)
        
        return driver
    
    def DefineDriver(driver_name):
        global driver
        if driver_name == "web":
            driver = Driver.StartWebdriver()
        elif driver_name == "talk":
            driver = Driver.StartTalkDriver()
        else:
            print("Driver should be defined as 'web' or 'talk'")
        
        return driver

class TestLink():
    global tls, tl_devkey
    
    tl_url = data["testlink"]["url"]
    tl_devkey = data["testlink"]["devkey"]
    tls = testlink.TestLinkHelper(tl_url, tl_devkey).connect(testlink.TestlinkAPIGeneric)

    def CreateTestPlan(project_name, new_testplan_name):
        print("deactive")
        '''testplan_file = open(Files.testplan, "w")
        testplan_file.write(new_testplan_name)
        testplan_file.close()

        tls.createTestPlan(new_testplan_name, project_name, note="", 
                            active=True, public=True, devKey=tl_devkey)'''

    def GetProject():
        print("deactive")
        '''project_name = dict(Logs.ReadUserInput())["project_name"]
        project = tls.getTestProjectByName(project_name)
        project_id = project["id"]

        return [project_id, project_name]'''

    def GetTestPlan():
        print("deactive")
        '''testplan_list = []
        
        project_name = TestLink.GetProject()[1]
        testplan_name = str(open(Files.testplan, "r").read())
        
        if bool(testplan_name) == True:
            testplan = tls.getTestPlanByName(project_name, testplan_name)
            testplan_id = testplan[0]["id"]
            
            testplan_list.append(testplan_id)
            testplan_list.append(testplan_name)

        return testplan_list'''
    
    def GetTestBuild(testplan_name):
        print("deactive")
        '''testplan_id = list(TestLink.GetTestPlan())[0]
        build_id = int(dict(tls.getLatestBuildForTestPlan(testplan_id))["id"])

        return build_id'''

    def Settings(testplan_name, build_name):
        print("deactive")
        '''user_input = Logs.ReadUserInput()

        project_name = str(user_input["project_name"])
        domain_name = str(user_input["domain_name"])
        user_id = str(user_input["user_id"])
        user_pw = str(user_input["user_pw"])
        version = str(user_input["version"])

        testplan_name = str(open(Files.testplan, "r").read())
        
        if bool(testplan_name) == True:
            testplan = tls.getTestPlanByName(project_name, testplan_name)
            testplan_id = testplan[0]["id"]
            
            project_id = TestLink.GetProject()[0]
            
            build_note = "Project: %s | Domain: %s | Account: %s/%s | Version: %s" % (project_name, domain_name, user_id, user_pw, version)
            tls.createBuild(testplan_id, build_name, buildnotes=build_note)'''

    def GetLastExecutionResult(testplan_id, tc_externalid):
        print("deactive")
        '''testcase = tls.getTestCase(testcaseexternalid=tc_externalid)

        testcase_name = testcase[0]["name"]

        division_id = int(testcase[0]["testsuite_id"])
        division = tls.getTestSuiteByID(testsuiteid=division_id)
        division_name = division["name"]
        #print(division_name)
        
        menu_id = int(division["parent_id"])
        menu = tls.getTestSuiteByID(testsuiteid=menu_id)
        menu_name = menu["name"]
        #print(menu_name)

        testcase_result = tls.getLastExecutionResult(testplanid=testplan_id, testcaseexternalid=tc_externalid)
        status = testcase_result[0]["status"]
        if status == "p":
            testcase_status = "Pass"
        elif status == "f":
            testcase_status = "Fail"
        elif status == "b":
            testcase_status = "Block"

        test_no = str(tc_externalid).split("-")[1]
        test_date = str(testcase_result[0]["execution_ts"]).split(" ")[0]
        #test_note = str(testcase_result[0]["notes"])
        
        testcase_execution = {
            "no": test_no,
            "menu": menu_name,
            "division": division_name,
            "name": testcase_name,
            "status": testcase_status,
            "test_date": test_date
        }

        return testcase_execution'''

    def Report(testcase_id, executed_status):
        print("deactive")
        '''domain_name = DefineCurrentURL().split("/ngw/app")[0]

        if executed_status == "Pass":
            tl_status = 'p'
            tl_note = "Test Case passes on domain [%s] at %s" % (domain_name, objects.date_time)
        else:
            tl_status = 'f'
            tl_note = "Test Case fails on domain [%s] at %s" % (domain_name, objects.date_time)
        
        testplan_id = TestLink.GetTestPlan()[0]
        testplan_name = TestLink.GetTestPlan()[1]
        build_id = TestLink.GetTestBuild(testplan_name)
        
        tls.reportTCResult(testplanid=testplan_id, testcaseexternalid=testcase_id, buildid=build_id, status=tl_status, notes=tl_note)'''

def PrintYellow(msg):
    '''• Usage: Color msg in yellow'''
    
    Logging('\033[93m' + str(msg) + '\033[0m')

    return msg

def PrintGreen(msg):
    '''• Usage: Color msg in green'''
    
    Logging('\033[92m' + str(msg) + '\033[0m')

    return msg

def PrintRed(msg):
    '''• Usage: Color msg in red'''
    
    Logging('\033[91m' + str(msg) + '\033[0m')

    return msg

class Waits():
    def WaitElementLoaded(time, xpath):
        '''• Usage: Wait until element VISIBLE in a selected time period'''
        
        WebDriverWait(driver, time).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element
    
    def Wait10s_ElementClickable(xpath):
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element

    def Wait10s_ElementLoaded(xpath):
        '''• Usage: Wait 10s until element VISIBLE'''
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element

    def WaitElementInvisibility(time, xpath):
        '''• Usage: Wait until element INVISIBLE in a selected time period'''
        
        WebDriverWait(driver, time).until(EC.invisibility_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)

        return element

    def Wait10s_ElementInvisibility(xpath):
        '''• Usage: Wait 10s until element INVISIBLE'''
        
        i=0
        for i in range(0,10):
            i+=1
            time.sleep(1)
            try:
                driver.find_element_by_xpath(xpath)
            except WebDriverException:
                break
    
    def WaitUntilPageIsLoaded(page_xpath):
        if bool(page_xpath) == True:
            # wait until page's element is present
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, page_xpath)))

        # check if the loading icon is not present at the page -> page is completely loaded
        try:
            WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH, "//div[@class='loading-dialog hide']")))
            FindPushNoti()
        except WebDriverException:
            pass

        '''If page_xpath=None/False -> only check if the loading icon is not present'''

class Commands():
    def FindElement(xpath):
        element = driver.find_element_by_xpath(xpath)

        return element
    
    def FindElement_ByCSS(css):
        element = driver.find_element_by_css_selector(css)

        return element

    def FindElements(xpath):
        element = driver.find_elements_by_xpath(xpath)

        return element

    def ClickElement(xpath):
        '''• Usage: Do the click on element
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        element.click()

        return element

    def ClickElements(xpath, element_position):
        '''• Usage: Do the click on element
                return WebElement'''

        element = driver.find_elements_by_xpath(xpath)
        time.sleep(1)
        element[element_position].click()

        return element

    def Wait10s_ClickElement(xpath):
        '''• Usage: Wait until the element visible and do the click
                return WebElement'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.click()

        return element

    def InputElement(xpath, value):
        '''• Usage: Send key value in input box
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        try:
            element.clear()
        except WebDriverException:
            pass
        element.send_keys(value)

        return element
    
    def InputElement_2Values(xpath, value1, value2):
        '''• Usage: Send key with 2 values in input box
                return WebElement'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        try:
            element.clear()
        except WebDriverException:
            pass
        element.send_keys(value1)
        element.send_keys(value2)

        return element

    def Wait10s_InputElement(xpath, value):
        '''• Usage: Wait until the input box visible and send key value
                return WebElement'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.send_keys(value)

        return element
    
    def SwitchToFrame(frame_xpath):
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, frame_xpath)))
        frame = Commands.FindElement(frame_xpath)
        driver.switch_to.frame(frame)

        return frame
    
    def SwitchToDefaultContent():
        driver.switch_to.default_content()

    def ScrollDown():
        '''• Usuage: Scroll down, default height (0,-301)'''
        
        driver.execute_script("window.scrollTo(0,300)")
    
    def ScrollUp():
        '''• Usuage: Scroll down, default height (300,0)'''
        
        driver.execute_script("window.scrollTo(301, 0)")
    
    def Selectbox_ByValue(xpath, value):
        '''• Usage: Wait until select box is loaded
                select by value, return select box
                value = str()'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        Select(element).select_by_value(value)

        return element
    
    def Selectbox_ByIndex(xpath, index_number):
        '''• Usage: Wait until select box is loaded
                select by the index, return select box
                index_number = int()'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        Select(element).select_by_index(index_number)

        return element
    
    def Selectbox_ByVisibleText(xpath, selected_text):
        '''• Usage: Wait until select box is loaded
                select by visible text, return select box
                visible text = str()'''

        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        Select(element).select_by_visible_text(selected_text)

        return element

    def MoveToElement(xpath):
        '''• Usage: Move to view element by ActionChains
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        actions = ActionChains(driver)
        actions.move_to_element(element)
        actions.perform()
        time.sleep(1)

        return element

    def NavigateTo(url):
        driver.get(url)

        return url
    
    def ExecuteScript(script):
        driver.execute_script(script)

    def ReloadBrowser(page_xpath):
        driver.refresh()
        Waits.WaitUntilPageIsLoaded(page_xpath)

    def SaveScreenShot(screenshot_location):
        driver.save_screenshot(screenshot_location)

    def ScrollIntoView(xpath):
        element = driver.find_element_by_xpath(xpath)
        driver.execute_script("arguments[0].scrollIntoView();", element)
        time.sleep(1)

        return element

    def ActionsWithContainedXpath(action, xpath, replaced_value):
        '''Usage: action name can be used (wait10s / click / find'''

        element_xpath = Functions.xpath_ConvertXpath(xpath, replaced_value)
        if action == "wait10s":
            Waits.Wait10s_ElementLoaded(element_xpath)
        elif action == "click":
            Commands.ClickElement(element_xpath)
        elif action == "find":
            Commands.FindElement(element_xpath)

class Functions():
    def GetElementText(xpath):
        '''• Usage: Get and return element_text as str()'''

        element_text = str(driver.find_element_by_xpath(xpath).text)

        return element_text
    
    def GetInputValue(xpath):
        '''• Usage: Get and return input_value as str()
                 Use this function if element is input box'''

        input_element = driver.find_element_by_xpath(xpath)
        input_value = str(input_element.get_attribute("value"))

        return input_value
    
    def GetElementAttribute(xpath, attribute):
        '''• Usage: Get and return element_attribute as str()
                        (attribute can be value of 'class', 'style'... '''

        element = driver.find_element_by_xpath(xpath)
        element_attribute = str(element.get_attribute(attribute))

        return element_attribute

    def GetListLength(xpath):
        '''• Usage: Count how many elements are visible
                return a number int()'''

        list_length = int(len(driver.find_elements_by_xpath(xpath)))

        return list_length
    
    def GetPageSource():
        page_source = driver.page_source
        
        return page_source
    
    def xpath_ConvertXpath(xpath, replaced_value):
        '''• Usage: xpath which is being used must be written in style 'replaced_text'
                return str()'''

        if type(replaced_value) == int():
            '''It's used to define the order number of element
                        E.g: xpath + "[" + str(i) + "]" '''
                        # i=int()
            element_xpath = str(xpath).replace("order_number", str(replaced_value))
        
        elif type(replaced_value) == str():
            ''' It's used to replace the text in xpath
                        E.g: xpath = xpath + [contains(., 'replaced_text')] '''
                        # replaced_text=str()
            element_xpath = str(xpath).replace("replaced_text", str(replaced_value))

        return element_xpath

    def getRandomNumber_fromSpecificRange(first_number, last_number):
        '''• Usage: Get a list of random numbers
                return a number int()'''

        random_number = int(random.randint(first_number, last_number))

        return random_number

    def getRandomList_fromSpecificRange(picked_numbers, assigned_range):
        '''• Usage: Get a list of random numbers and remove duplicated number
                return a list()'''

        random_number = random(randint(range(assigned_range)))

        random_list = []
        i=1
        for i in range(assigned_range):
            random_number = random(randint(range(assigned_range)))
            random_list.append(random_number)
            
            random_list = list(dict.fromkeys(random_list))
            if len(random_list) == picked_numbers:
                break
            
            i+=1 

        return random_list

    def RemoveDuplicate_fromList(selected_list):
        '''• Usage: Remove duplicated items in the assigned list
                return the assigned list without duplicated item'''
        
        selected_list = list(dict.fromkeys(selected_list))

        return selected_list

    def checkIf_ElementVisible(xpath):
        '''• Usage: check element is visible
                    return True if element is visible'''
        
        try:
            driver.find_element_by_xpath(xpath)
            return True
        except WebDriverException:
            return False

    def waitIf_ElementVisible(xpath):
        '''• Usage: Wait 10s until element is visible
                    return True if element is visible'''
        
        try:
            Waits.Wait10s_ElementLoaded(xpath)
            return True
        except WebDriverException:
            return False

class Logs():
    def CreateLogFiles():
        date_time = datetime.now().strftime("%Y/%m/%d, %H:%M:%S")
        section_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
        logs = [Files.execution_log, Files.fail_log, Files.error_log, Files.testcase_log]
        for log in logs:
            if ".txt" in str(log):
                open(log, "w").close()
            
            elif log == Files.testcase_log:
                wb = load_workbook(Files.testcase_log)
                ws = wb.active

                last_row = ws.max_row
                for row_number in range(2,last_row):
                    # date_cell_value = ws.cell(row=row_number, column=6).value
                    
                    # if date_cell_value != str(objects.today):
                    #     ws.cell(row=row_number, column=5).value = ""
                    #     ws.cell(row=row_number, column=6).value = ""
                    #     ws.cell(row=row_number, column=7).value = ""
                    
                    for column in range(4, 15):
                        column +=1
                        ws.cell(row=row_number, column=column).value = ""
                    
                    row_number+=1

                ws.cell(row=2, column=9).value = section_id
                
                wb.save(Files.testcase_log)
        
        return {"logs": logs, "section_id": section_id}

    def CreateTestCaseFiles(section_id):
        testcase_filename = "%s_result_%s.xlsx" % (Files.testplan_name, section_id)
        testcase_file  = Files.test_log_folder + testcase_filename

        wb = load_workbook(testcase_file)
        ws = wb.active

        last_row = ws.max_row
        for row_number in range(2,last_row):
            for column in range(4, 15):
                column +=1
                ws.cell(row=row_number, column=column).value = ""
            
            row_number+=1
        
        wb.save(testcase_file)

    def TestCaseStatus(testcase_no, status):
        wb = load_workbook(Files.testcase_log)
        ws = wb.active

        row_update = testcase_no-1

        ws.cell(row=row_update, column=5).value = status
        ws.cell(row=row_update, column=6).value = str(objects.today)

        wb.save(Files.testcase_log)
    
    def ClearExcel_CollectMenu():
        # Clear data in excel file - Write menu name in file
        wb = load_workbook(Files.collect_menu_log)
        ws = wb.active

        menu_list = dict(data["excel_menu_list"])
        
        last_row = ws.max_row
        if last_row > 1:
            ws.delete_rows(2, last_row)
            for menu in menu_list.keys():
                current_row = menu_list[menu]
                ws.cell(row=current_row, column=1).value = menu
        
        ws.cell(row=2, column=5).value = "Groupware" # project_name

        wb.save(Files.collect_menu_log)
    
    def Value_CollectMenu(row, column):
        # Collect cell value in excel file
        wb = load_workbook(Files.collect_menu_log)
        ws = wb.active

        cell_value = ws.cell(row=row, column=column).value
        wb.save(Files.collect_menu_log)

        return cell_value

    def UpdateSelection_CollectMenu(selected_menu):
        # Update cell_value = True from selected_menu list in excel file 
        wb = load_workbook(Files.collect_menu_log)
        ws = wb.active

        last_row = ws.max_row
    
        for row_number in range(2,last_row):
            menu_cell = ws.cell(row=row_number, column=1).value
            
            if menu_cell in selected_menu:
                ws.cell(row=row_number, column=2).value = True
        
        wb.save(Files.collect_menu_log)

    def UpdateSuccess_ColectMenu(menu):
        # Write result = True if test execution finished without error (groupware_execution.py)

        wb = load_workbook(Files.collect_menu_log)
        ws = wb.active
        
        menu_list = dict(data["excel_menu_list"])
        current_row = menu_list[menu]
                
        ws.cell(row=current_row, column=3).value = True
        wb.save(Files.collect_menu_log)
    
    def CheckResult_CollectMenu():
        # Collect list including execution which failed while running
        
        wb = load_workbook(Files.collect_menu_log)
        ws = wb.active
        
        error_list = []
        menu_list = dict(data["excel_menu_list"])
        
        for menu in menu_list.keys():
            current_row = menu_list[menu]
            
            selected_value = ws.cell(row=current_row, column=2).value
            result_value = ws.cell(row=current_row, column=3).value
            
            if selected_value != result_value:
                error_list.append(menu)
                # if execution finish without error, 
                # selected_value = result_value = True

        wb.save(Files.collect_menu_log)

        return error_list
    
    def UserInput(**user_input):
        # Write user input including (data row and column are fixed):
        #       project_name, domain_name, user_id, user_pw, version

        domain_name = user_input["domain_name"]
        user_id = user_input["user_id"]
        user_pw = user_input["user_pw"]
        date = user_input["date"]

        wb = load_workbook(Files.section_history)
        ws = wb.active
        
        ws.cell(row=2, column=3).value = domain_name
        ws.cell(row=2, column=4).value = user_id
        ws.cell(row=2, column=5).value = user_pw
        ws.cell(row=2, column=7).value = date

        wb.save(Files.section_history)

    def ReadUserInput():
        wb = load_workbook(Files.section_history)
        ws = wb.active
        
        user_input = {}

        user_input["project_name"] = ws.cell(row=2, column=2).value
        user_input["domain_name"] = ws.cell(row=2, column=3).value
        user_input["user_id"] = ws.cell(row=2, column=4).value
        user_input["user_pw"] = ws.cell(row=2, column=5).value
        user_input["version"] = ws.cell(row=2, column=6).value
        user_input["date"] = ws.cell(row=2, column=7).value

        wb.save(Files.collect_menu_log)

        return user_input

def Logging(*messages):
    msg = str(" ".join(list(messages))) 
    print(msg)
    log_msg = open(Files.execution_log, "a")
    written_msg = str(msg).encode(encoding="ascii",errors="ignore")
    log_msg.write(str(written_msg, 'utf-8') + "\n")
    log_msg.close()

def TestCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging(description)
    
    if status == "Pass":
        Logging(objects.testcase_pass)
    else:
        Logging(objects.testcase_fail)

    wb = openpyxl.load_workbook(Files.testcase_log)
    current_sheet = wb.active
        
    row_length = len(list(current_sheet.rows)) + 1
    for current_row in range(1, row_length):
        menu_cell = current_sheet.cell(row=current_row, column=2).value
        testcase_cell = current_sheet.cell(row=current_row, column=4).value
        
        if current_row > 1:
            current_sheet.cell(row=current_row, column=1).value = str(current_row-1)

        if testcase_cell == testcase and menu_cell == menu:
            current_sheet.cell(row=current_row, column=6).value = "%s %s" % (objects.time1, objects.today)
            
            if status == "Pass":
                Logging(objects.testcase_pass)
                current_sheet.cell(row=current_row, column=5).value = "Pass"
            else:
                Logging(objects.testcase_fail)
                current_sheet.cell(row=current_row, column=5).value = "Fail"
            
            current_sheet.cell(row=current_row, column=7).value = Files.testplan_name

    wb.save(Files.testcase_log)

def OldTestCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging(description)
    
    if status == "Pass":
        Logging(objects.testcase_pass)
    else:
        Logging(objects.testcase_fail)
    
    wb = openpyxl.load_workbook(Files.testcase_log)
    current_sheet = wb.active
        
    row_length = len(list(current_sheet.rows)) + 1
    for current_row in range(1, row_length):
        menu_cell = current_sheet.cell(row=current_row, column=1).value
        testcase_cell = current_sheet.cell(row=current_row, column=3).value
        if testcase_cell == testcase and menu_cell == menu:
            current_sheet.cell(row=current_row, column=6).value = "test date"
            if status == "Pass":
                Logging(objects.testcase_pass)
                current_sheet.cell(row=current_row, column=4).value = "Pass"
                current_sheet.cell(row=current_row, column=5).value = ""
            else:
                Logging(objects.testcase_fail)
                current_sheet.cell(row=current_row, column=4).value = "Fail"
                current_sheet.cell(row=current_row, column=5).value = description  

    wb.save(Files.testcase_log)

def DefineCurrentTime():
    now = datetime.now()
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    
    return date_time

def DefineCurrentURL():
    current_url = str(driver.current_url)

    return current_url

def CollectListData(list_footer, page_total):
    Waits.Wait10s_ElementLoaded(list_footer)
    list_footer = Functions.GetElementText(list_footer)
    total_items = int(list_footer.replace(",", "").split(" ")[1])

    try:
        Waits.WaitElementLoaded(3, page_total)
        total_pages = int(Functions.GetElementText(page_total))
    except WebDriverException:
        try:
            Commands.FindElement("//div[@class='load-more']/button/span")
            total_pages = 2
        except WebDriverException:
            total_pages = 0

    list_data = {
        "total_items": total_items,
        "total_pages": total_pages
    }

    return list_data

def TestLinkResult_Pass(external_id):
    Logging(objects.testcase_pass)

def TestLinkResult_Fail(external_id):
    Logging(objects.testcase_fail)

def ValidateFailResultAndSystem(fail_msg):
    Logging(fail_msg)
    append_fail_result = open(Files.fail_log, "a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()

def wrapper(function, args):
    outcome = function(*args)

    return outcome

def AccessGroupwareMenu(name, page_xpath):
    Waits.WaitUntilPageIsLoaded(None)

    Waits.Wait10s_ElementLoaded("//*[@id='main-navi']/nav-menu-react/nav/a[@data-name='" + name + "']")

    menu = Commands.FindElement("//*[@id='main-navi']/nav-menu-react/nav/a[@data-name='" + name + "']")
    menu_display = menu.get_attribute("style")
    if menu_display == 'display: none;':
        FindPushNoti()

        Commands.MoveToElement("//a[@class='lm--navDropdown']")
        Logging("Hover nav dropdown (more menu)")

        Waits.WaitElementLoaded(5, "//ul[@class='ulnavDropdown-child']/li/a[@data-name='" + name + "']")
        Commands.ClickElement("//a[@data-name='" + name + "' and @style='display: inline;']")
        Logging("Access menu from dropdown menu")
    else:
        Waits.WaitUntilPageIsLoaded("//*[@id='main-navi']/nav-menu-react/nav/a[@data-name='" + name + "']")
        menu.click()
    
    try:
        if name == "nhr,HR":
            Waits.WaitElementLoaded(20, "//*[@id='nhrIframe']")
            Commands.SwitchToFrame("//*[@id='nhrIframe']")
            Waits.WaitElementLoaded(20, "//div[contains(@class, 'dashboard-wrapper')]")
            Commands.SwitchToDefaultContent()
        else:
            Waits.Wait10s_ElementLoaded(page_xpath)
            time.sleep(1)
     
        Logging("Access menu [" + name + "] successfully" +  objects.testcase_pass)
        access_result = True
    except WebDriverException:
        Logging("Fail to access menu [" + name + "] " + objects.testcase_fail)
        access_result = False

    time.sleep(1)

    if name == "resource,Resource":
        time.sleep(2)
    
    Waits.WaitUntilPageIsLoaded(None)

    return access_result

def Timecard_ExceptedUserValidation():
    try:
        try:
            Waits.WaitElementLoaded(5, "//div[@class='timecard']/div[3]/div/button")
            clock_button = Commands.FindElement("//div[@class='timecard']/div[3]/div/button")
        except WebDriverException:
            clock_button = Commands.FindElement("//div[@class='timecard']/div[2]/div/button")
        finally:
            Logging("Wait until clock in/out is visible")
            excepted_user = False
            clock_button_text = str(clock_button.text)
    except WebDriverException:
        Commands.FindElement("//div[@ng-if='timecard.serviceExcepted()']")
        clock_button_text = "Excepted User"
        excepted_user = True
    
    tc_data = {
        "excepted_user": excepted_user,
        "clock_button_text": clock_button_text
    }

    return tc_data

def Timecard_ConfirmWorkingTime(clock_button_text):
    if "Clock In" in clock_button_text:
        Waits.WaitElementLoaded(5, "//button[contains(., 'Clock In')]")
        Commands.Wait10s_ClickElement("//button[contains(., 'Clock In')]")
        Logging("Clock In timecard")
    
        '''try:
            Waits.WaitElementLoaded(5, "//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
            Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
            Logging("Cancel late modal ")

            Waits.WaitElementLoaded(5, "//button[contains(., 'Clock Out')]")
            Commands.Wait10s_ClickElement("//button[contains(., 'Clock Out')]")
            Logging("Clock Out timecard")
            
            Waits.WaitElementLoaded(5, "//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
            Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
            Logging("Cancel Leave Early")

            TestCase_LogResult(**data["testcase_result"]["right_sidebar"]["timecard"]["pass"])
        except WebDriverException:
            Commands.ClickElement("//h4[contains(., 'Error')]/preceding-sibling::button")
            TestCase_LogResult(**data["testcase_result"]["right_sidebar"]["timecard"]["fail"])'''
        Waits.WaitElementLoaded(5, "//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
        Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
        Logging("Cancel late modal ")

        Waits.WaitElementLoaded(5, "//button[contains(., 'Clock Out')]")
        Commands.Wait10s_ClickElement("//button[contains(., 'Clock Out')]")
        Logging("Clock Out timecard")
        
        Waits.WaitElementLoaded(5, "//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
        Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
        Logging("Cancel Leave Early")
    elif "Clock Out" in clock_button_text:
        Waits.WaitElementLoaded(5, "//button[contains(., 'Clock Out')]")
        Logging("Clock Out timecard")

        ValidateModalDialog_ErrorAlert()

        '''try:
            Waits.WaitElementLoaded(5, "//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
            Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
            Logging("Cancel Leave Early")
            TestCase_LogResult(**data["testcase_result"]["right_sidebar"]["timecard"]["pass"])
        except WebDriverException:
            Commands.ClickElement("//h4[contains(., 'Error')]/preceding-sibling::button")
            TestCase_LogResult(**data["testcase_result"]["right_sidebar"]["timecard"]["fail"])'''
        Waits.WaitElementLoaded(5, "//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
        Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard')]")
        Logging("Cancel Leave Early")
    elif "Continue Work" in clock_button_text:
        Logging("User already punched in/out")

        time.sleep(2)
        
        Commands.ClickElement("//button[@data-ng-click='timecard.continueWork(false)']")
        Logging("Click Continue Work")

        Commands.Wait10s_ClickElement("//button[contains(., 'Clock Out')]")
        Logging("Click Clock out button")

        Commands.Wait10s_ClickElement("//button[@ng-click='timecard.clockOutHandlerAgain($event)']")
        Logging("Click confirm clock out button")

        Commands.Wait10s_ClickElement("//button[contains(., 'Cancel') and contains(@ng-click, 'timecard.confirmClockOutEarly = false')]")
        Logging("Cancel Leave Early")
        
        Waits.WaitElementLoaded(5, "//button[@data-ng-click='timecard.continueWork(false)']")
        Logging("Wait for Continue Work button to be visible")

        TestCase_LogResult(**data["testcase_result"]["right_sidebar"]["timecard"]["pass"])

def Groupware_CheckWorkingTimeInSidebar(driver):
    Commands.ClickElement("//a[contains(@class, 'open-sidebar')]")
    Logging("Open Right Sidebar")

    try:
        Waits.Wait10s_ElementLoaded("//div[@data-ng-include='templateTimecardPath']")
        Logging("Timecard modal is visible")
        timecard_modal = True
    except WebDriverException:
        timecard_modal = False
    
    if timecard_modal == True:
        tc_data = Timecard_ExceptedUserValidation()
        excepted_user = tc_data["excepted_user"]
        clock_button_text = tc_data["clock_button_text"]
        if excepted_user == False:
            Logging("Confirm working time in timecard modal")
            Timecard_ConfirmWorkingTime(clock_button_text)
        else:
            Logging("Current user is timecard excepted user")
    else:
        ValidateFailResultAndSystem("Cannot find user timecard information")
        Logging(objects.testcase_fail)

def CommonWriteItem(red_pen, title_target, mail_content):
    try:
        Waits.Wait10s_ElementLoaded(data["common"]["loading_dialog"])
        FindPushNoti()
    except WebDriverException:
        pass
    
    Waits.WaitElementLoaded(15, red_pen)
    Commands.Wait10s_ClickElement(red_pen)
    Logging("Write - Click Create button")

    Waits.WaitElementLoaded(15, "//*[@class='tox-edit-area__iframe']")
    Waits.WaitUntilPageIsLoaded(None)

    title = Commands.InputElement(title_target, objects.hanbiro_title)
    Logging("Write - Input title / subject")
    Logging(">>> Title: [" + title.get_attribute("value") + "] is input")
    
    Commands.ScrollDown()
    Waits.Wait10s_ElementLoaded("//*[@class='tox-edit-area__iframe']")
    time.sleep(1)
    
    Commands.SwitchToFrame("//*[@class='tox-edit-area__iframe']")
    time.sleep(1)

    try:
        Commands.InputElement("//*[@id='tinymce']/p", mail_content)
        Logging("Write - Content is empty - Input content")
    except WebDriverException:
        try:
            Commands.InputElement("//*[@id='tinymce']/div", mail_content)
            Logging("Write - Mail - Content is empty - Input content")
        except WebDriverException:
                Commands.InputElement("//*[@id='tinymce']/div/div", mail_content)
                Logging("Write - Mail - Input content")

    Commands.SwitchToDefaultContent()
    
    CloseAutosave()

def InputContent():
    Waits.Wait10s_ElementLoaded("//*[@class='tox-edit-area__iframe']")
    time.sleep(1)
    Commands.SwitchToFrame("//*[@class='tox-edit-area__iframe']")
    Waits.Wait10s_ElementLoaded("//*[@id='tinymce']/p")
    Commands.InputElement("//*[@id='tinymce']/p", objects.hanbiro_content)
    time.sleep(1)
    comment = Functions.GetElementText("//*[@id='tinymce']/p")
    Commands.SwitchToDefaultContent()
    
    return comment

def EditContent():
    Commands.SwitchToFrame("//*[@class='tox-edit-area__iframe']")
    Waits.Wait10s_ElementLoaded("//*[@id='tinymce']/p")
    time.sleep(1)
    content = Commands.FindElement("//*[@id='tinymce']/p")
    comment = content.text
    content.clear()
    content.send_keys(comment + Keys.ENTER + objects.content_edit)
    Commands.SwitchToDefaultContent()

def FindPushNoti():
    try:
        push_no = int(len(driver.find_elements_by_css_selector(".ui-pnotify-closer > .fa"))) + 1
        for i in range (1, push_no):
            driver.find_element_by_css_selector(".ui-pnotify-closer > .fa").click()
            Logging("Write - Close push notification " + str(i))
    except WebDriverException:
        pass

def SelectboxOption(select_target):
    # Desc: Select random option from select box
    Commands.ClickElement(select_target)
    option_list = Functions.GetListLength(select_target + "/option")
    if option_list > 1:
        option_range = option_list
        selectbox = Select(Waits.Wait10s_ElementLoaded(select_target))
        selectbox.select_by_index(int(random.randint(1, option_range)))
        select_name = Commands.FindElement(select_target + "/option[contains(@selected, 'selected')]")
        Logging("Option [" + select_name.text + "] is selected")
    else:
        Logging("No option to select")

def ValidateListTotal(list_footer):
    Waits.Wait10s_ElementLoaded(list_footer)
    total_counter = Functions.GetElementText(list_footer)
    total_counter_number = int(total_counter.replace(",", "").split(" ")[1])
    Logging("!!! MESSAGE: Total list number: " + total_counter)

    return total_counter_number

def ValidateListNumberUpdate(total_counter_number, total_counter_update_number):
    # Desc: Check api by comparing the list counter
    if total_counter_number < total_counter_update_number:
        Logging("Issue is saved successfully -> API passed")
        Logging(objects.testcase_pass)
        list_result = True
    else:
        Logging("Fail to save issue -> API failed")
        Logging(objects.testcase_fail)
        list_result = False
    
    return list_result

def validate_user_list(selected_list_target, selected_user_pre_target, selected_user_suf_target, save_target):
    # Select user lisr randomly from Organization by selecting the last department
    selected_list = Functions.GetListLength(selected_list_target)
    selected_list_range = selected_list + 1 
    selected_user = None   

    for user in range(1, selected_list_range):
        selected_user_xpath = selected_user_pre_target + str(user) + selected_user_suf_target
        selected_user = Commands.FindElement(selected_user_xpath)
        Logging("selected_user: " + selected_user)
        
    Commands.ClickElement(save_target)
    Logging("Save Organization tree")

    return selected_user    

def CloseAutosave():
    try:
        Commands.FindElement("//h4[contains(.,'Automatically saved file found!')]")
        try:
            Commands.ClickElement("//*[@title='Remove All']")
            Logging("Autosave - Autosave content is completely deleted")
        except WebDriverException:
            Commands.ClickElement("//*[@title='Close']")
            Logging("Autosave - Close autosave successfully")
    except WebDriverException:
        Logging("Autosave - No autosave content found")

def Org_SelectUser(org_tree, org_input, org_plus, org_save, recipient_id):
    Commands.Wait10s_ClickElement(org_tree)
    Logging("Organization Tree - Click Org tree button")
    
    org_input_element = Waits.Wait10s_ElementLoaded(org_input)
    time.sleep(1)

    Commands.ClickElement(org_input)
    Commands.InputElement_2Values(org_input, recipient_id, Keys.RETURN)
    Logging("Input recipient key in search box")
    
    Waits.WaitUntilPageIsLoaded(None)
    Waits.Wait10s_ElementLoaded(expense_dict["user_org_xpath"])

    try:
        Commands.FindElement(expense_dict["recipient_xpath"])
        recipients = Commands.FindElements(expense_dict["recipient_xpath"])
    except WebDriverException:
        recipients = Commands.FindElements(expense_dict["recipient_block_xpath"])
    recipients[0].click()
    Logging("Select recipient")

    Commands.ClickElement(org_plus)
    Logging("Add recipient")

    Commands.ClickElement(org_save)
    Logging(" Save selected recipient")

    selected = True

    return selected

def DefineItemPosition(list_number):
    Waits.Wait10s_ElementLoaded(list_number)
    time.sleep(1)

    # Count how many items found in the list
    item_number = Functions.GetListLength(list_number)
    Logging("---------- Define Itemt Position - List Number: " + str(item_number))

    '''
    full item xpath target: "//div[@id='message-list']/react-mail-list/div/div[x]/div/div/a/span[2]" (x = item_position)
    x = 1: the first item
    x = 2: the second item
    '''

    item_position = random.randint(1, item_number)
    if "clouddisk-list" in list_number:
        item_position = random.randint(2, item_number)
    else:
        item_position = random.randint(1, item_number)
    Logging("---------- Define Itemt Position - Item_position: " + str(item_position))

    return item_position

def FindSecureItem(pre_secure, item_position, suf_secure):
    try:
        secure_item = Commands.FindElement(pre_secure + str(item_position) + suf_secure)
        Logging("View Content - Secure item is found")
    
        secure_item.click()
        Logging("View Secure Content - Click on secure item")

        Commands.SwitchToFrame("//*[starts-with(@id, 'iframeTypePassword')]")

        Waits.WaitElementLoaded(5, "//input[@type='password']")
        Commands.InputElement_2Values("//input[@type='password']", data["security_password"], Keys.RETURN)
        Logging("View Secure Content - Input security password")

        Commands.SwitchToDefaultContent()
    except WebDriverException:
        pass

def ViewContent(list_footer, list_number, pre_target, suf_target, pre_secure, suf_secure, content_p):
    Waits.Wait10s_ElementLoaded(list_number)

    item_position = DefineItemPosition(list_number)
    Logging("item_position " + str(item_position))

    item_position1 = DefineItemPosition(list_number)
    Logging("item_position1 " + str(item_position1))
    Logging("View Content - Define item position")
    
    item = Commands.FindElement(pre_target + str(item_position) + suf_target)
    try: 
        item.click()
        item_text = item.text
        Logging("View Normal Content - Click on item")
    except WebDriverException:
        item_position = item_position1
        item.click()
        Logging("View Normal Content - Click on item")
    
    try:
        Waits.Wait10s_ElementLoaded("//*[@class='tox-edit-area__iframe']")
    except WebDriverException:
        Waits.Wait10s_ElementLoaded(content_p)
    
    if item_text.strip() in driver.page_source:
        Logging("View Content - Preview content successfully")
        Logging(objects.testcase_pass)
        result = True
    else:
        Logging("View Content - Fail to preview content")
        Logging(objects.testcase_fail)
        result = False
    
    time.sleep(1)

    return result

def searchInput(list_footer, list_number, pre_target, suf_target, search_input_xpath, result_target):
    '''Input key word to search input and enter key word for searching'''
    
    Waits.Wait10s_ElementLoaded(list_number)
    time.sleep(1)
    
    item_position = DefineItemPosition(list_number)
    
    try:
        item = Commands.FindElement(pre_target + str(item_position) + suf_target)
    except WebDriverException:
        item_position = item_position + 1
        item = Commands.FindElement(pre_target + str(item_position) + suf_target)

    selected_item = str(item.text).strip()

    Waits.Wait10s_ElementLoaded(search_input_xpath)
    Commands.InputElement_2Values(search_input_xpath, selected_item, Keys.ENTER)
    Logging("Search Input - Enter key words")

    search_keys = Functions.GetInputValue(search_input_xpath)
    Logging("Search Input - Key word: " + search_keys)

    time.sleep(2)

    Waits.Wait10s_ElementLoaded(result_target)
    result = Functions.GetElementText(result_target)
    Logging("Search Input - The result found: " + result)
    result_text = result.strip()
    replaced_result = result.replace(result_text, search_keys)
    search_item = Commands.FindElement(result_target + "[contains(., '" + replaced_result + "')]")
    
    try:
        search_item
        Logging("Search Input - Search successfully")
        Logging(objects.testcase_pass)
        
        Commands.InputElement(search_input_xpath, Keys.RETURN)
        Waits.WaitUntilPageIsLoaded(list_footer)

        search_result = True
    except WebDriverException:
        Logging("Search Input - Fail to search")
        Logging(objects.testcase_fail)
        search_result = False

    return search_result

def Notification_CheckReceive(hanbiro_title, leftside_badge, menu_badge):
    # Desc: Verify notification pop-up incoming
    try:
        Waits.WaitElementLoaded(20, "//a[contains(., '%s')]" % str(hanbiro_title).lower())
        Logging("Notification - Sent successfully")
        Counter_CheckReceive(leftside_badge, menu_badge)
    except WebDriverException:
        Logging("Notification - Notification not found")
        Logging(objects.testcase_fail)

def Counter_CheckReceive(leftside_badge, menu_badge):
    # Compare unread counter of sub menu from left sidebar and menu top counter
    leftside_counter = Functions.GetElementText(leftside_badge)
    menu_counter = Functions.GetElementText(menu_badge)
    if leftside_counter.replace("," , "") == menu_counter:
        Logging("Counter - Top and left menu counter are same")
    else:
        Logging("Counter - Top and left menu counter are not same")

def Autocomplete_SelectRecipient(hanbiro_user, address_holder):
    # Select user from autocomplete by the login id
    Commands.Wait10s_ClickElement(address_holder)
    Commands.InputElement("//*[@class='ui-autocomplete-input']", hanbiro_user)
    Waits.Wait10s_ElementLoaded("//html/body/ul/li/div[1]")
    recipient = Commands.FindElement("//html/body/ul/li/div[1]")
    Logging("Autocomplete - Recipient: " + recipient.text)
    time.sleep(1)
    recipient.click()
    Waits.WaitElementLoaded(3, data["mail"]["recipient_tag"])

def Attachment_SelectPC():
    Commands.ScrollDown()
    Commands.ClickElement(data["attachment"]["attach_button"])
    Logging("PC Attachment - Click Attach file button")

    Commands.ScrollDown()
    Commands.InputElement(data["attachment"]["pc_container"], Files.image_attachment)
    Logging("PC Attachment - Collect file from local folder")

    Commands.ScrollDown()
    pc_attach_name = Functions.GetElementText(data["attachment"]["file_placeholder"])
    
    return pc_attach_name

def Attachment_SelectImageCloudDisk():
    '''Search an image with extension .jpg and select as clouddisk attachment'''

    # Access CloudDisk tab from Attachment
    Commands.ClickElement(data["attachment"]["clouddisk_upload"])

    # Click to add clouddisk file
    Commands.Wait10s_ClickElement(data["attachment"]["add_clouddisk"])

    file = 0
    try:
        Waits.Wait10s_ElementLoaded(data["attachment"]["clouddisk_file"])
        FindPushNoti()

        Commands.InputElement_2Values(data["attachment"]["cloud_container"], ".png", Keys.RETURN)
        Logging("CloudDisk Attachment - Search file extension")

        time.sleep(2)

        try:
            Commands.Wait10s_ClickElement(data["attachment"]["clouddisk_file"])
            Logging("CloudDisk Attachment - Select image from searching result")
            file = 1
        except WebDriverException:
            Logging("Cannot find clouddisk file")

        if file > 0:
            # After file is selected successfully,
            # the counter number from "Selected File" will be updated
            # "1" = 1 file is selected
            select_counter = Functions.GetElementText(data["attachment"]["counter_selected"])
            if select_counter == "1":
                Commands.ClickElement(data["attachment"]["attach_cloud"])
            else:
                Logging("CloudDisk Attachment - Cannot click attach file button")
    except WebDriverException:
        pass
    
    # Validation if selected file is displayed in placeholder
    if file > 0:
        Waits.Wait10s_ElementLoaded(data["attachment"]["cloud_placeholder"])
        Logging("CloudDisk Attachment - Select clouddisk file successfully")
        cloud_attach_name = Functions.GetElementText(data["attachment"]["cloud_placeholder"])
    else:
        Logging("CloudDisk Attachment - Fail to select clouddisk file")
        cloud_attach_name = None
        Commands.ClickElement(data["attachment"]["close_clouddisk"])
        time.sleep(1)

    return cloud_attach_name

def DownloadFiles(domain_name, page_element):
    total_file1 = int(len(os.listdir(Files.download_folder)))
    Logging("total_file1: " + total_file1)

    page_element = Commands.FindElement(page_element)
    Logging("Define page's element")

    Commands.ClickElement("//a[@title='Download']")
    Logging("Click Download button")

    time.sleep(2)

    current_url = driver.current_url
    menu = current_url.split(domain_name + "/")[1].split("/")[0]

    total_file2 = int(len(os.listdir(Files.download_folder)))
    Logging("total_file2: " + total_file2)

    if total_file2 > total_file1:
        Logging("[" + menu + "]" + " File is downloaded successfully")
        Logging(objects.testcase_pass)
    else:
        Logging("[" + menu + "]" + " Fail to download file")
        Logging(objects.testcase_fail)
        try:
            page_element
            Logging("Page is not navigated")
        except WebDriverException:
            Logging("Page is navigated")
            Commands.ReloadBrowser(page_element)

def ValidateModalDialog_ErrorAlert():
    try:
        Waits.WaitElementLoaded(2, "//div[@class='modal-dialog ui-draggable']/div/div[contains(@class, 'modal-body')]")
        msg = Functions.GetElementText("//div[@class='modal-dialog ui-draggable']/div/div[contains(@class, 'modal-body')]")
        Logging("Unexpected alert: " + msg)
        time.sleep(1)
        Commands.ClickElement("//div[@class='modal-dialog ui-draggable']/div/div[contains(@class, 'modal-header')]/button")
        alert_msg = True
    except WebDriverException:
        alert_msg = None

    return alert_msg

def TCResult_ValidateAlertMsg(menu, testcase, msg):
    alert_msg = ValidateModalDialog_ErrorAlert()
    if bool(alert_msg) == True:
        data["testcase_result"]["" + menu + ""]["" + testcase + ""]["fail"].update({"description": "Fail to " + msg + " with alert" + str(alert_msg)})

def Counter_CheckCounterNumber(top_counter, left_counter):
    try:
        top_counter = Commands.FindElement(top_counter)
        topcounter_number = top_counter.text
        Logging("Top counter: " + topcounter_number)
    except WebDriverException:
        topcounter_number = 0
        Logging("No Top counter")
    
    try:
        leftcounter_number = Functions.GetElementText(left_counter)
        Logging("Left counter" + str(topcounter_number))
    except WebDriverException:
        leftcounter_number = 0
        Logging("No Left counter")

    counter_number = [topcounter_number, leftcounter_number]

    return counter_number

def Notification_ValidateIncomingNotification():
    menu_badge = Functions.GetListLength("//span[@class='badge']")
    if menu_badge > 1:
        Logging("Common Notification - Notification is valid")
        Logging(objects.testcase_pass)
        notification = True
    else:
        Logging("Common Notification -  Notification is invalid")
        Logging(objects.testcase_fail)
        notification = False

    return notification

def List_ColectItemList(list_target, item_suf):
    list_number = Functions.GetListLength(list_target)

    item_list = []

    range1 = range(1, list_number)
    range2 = range(2, list_number - 1)

    try:
        list_length = range1
    except WebDriverException:
        list_length = range2
    
    item_name = {
        "common1": ".message-item .summary .text",
        "common2": ".message-item .summary",
        "project_work": "table.tree-grid .work-name",
        "contact": ".contacts-list .user",
        "expense": ".message-item .text"
    }

    url = str(driver.current_url)
    if "addrbook" in url:
        css_item_name = item_name["contact"]
    elif "project" in url:
        css_item_name = item_name["project_work"]
    elif "clouddisk" in url:
        css_item_name = item_name["common2"]
    elif "asset" in url:
        css_item_name = item_name["common2"]
    elif "expense" in url:
        css_item_name = item_name["expense"]
    else:
        css_item_name = item_name["common1"]

    for x in list_length:
        msg_list = Commands.FindElement(list_target + "[" + str(x) + "]")
        item = msg_list.find_element_by_css_selector(css_item_name)
        item_text = item.text
        item_list.append(item_text)
    
    return item_list

def List_ValidateListMovingPage(list_target, item_suf, page_total_xpath, nextpage_icon):
    Waits.Wait10s_ElementLoaded(list_target)

    list1 = List_ColectItemList(list_target, item_suf)
    Logging("Collect list 1")

    Waits.Wait10s_ElementLoaded(data["common"]["loading_dialog"])

    driver.execute_script("window.scrollTo(0,800)")
    
    time.sleep(2)
    
    try:
        Commands.FindElement("//div[@class='load-more']/button/span")
        # list_type = "Load More"
        page_number = 2
    except WebDriverException:
        try:
            # list_type = "Load Page"
            page_number = int(Functions.GetElementText(page_total_xpath))
            Logging("page_number" + str(page_number))
        except WebDriverException:
            # list_type = "Load More"
            page_number = 0
    
    if page_number > 1:
        try:
            next_page = Commands.FindElements("//span[contains(@class, 'next_paging')]")
        except WebDriverException:
            next_page = Commands.FindElements("//div[@class='load-more']/button/span")
        finally:
            time.sleep(1)
            try:
                next_page[0].click()
            except WebDriverException:
                #next_page[1].click()
                i=0
                next_page_list = len(next_page)
                for i in range(0, next_page_list):
                    i+=1
                    try:
                        next_page[i].click()
                        break
                    except WebDriverException:
                        pass
            finally:
                Logging("Click next page icon in list")

        time.sleep(1)
        list2 = List_ColectItemList(list_target, item_suf)
        if list2 != list1:
            Logging("Access next page - Collect list 2")
            result = True
        else:
            result = False
        
        if bool(result) == True:
            Logging("Validate list next page successfully")
            Logging(objects.testcase_pass)
        else:
            Logging("List does not change when moving page")
            Logging(objects.testcase_fail)
    else:
        Logging("Cannot move page due to number of page total")
        result = None
    
    return result
    
def CopyArchive_ValidateDocumentTransfer(domain_name, folder, folder_name, document_name):
    current_url = DefineCurrentURL()
    if "/archive/" not in current_url:
        driver.get(domain_name + "/archive/search/detail/")
        time.sleep(1)
        
        Commands.ClickElement(data["archive"]["nav_archive"])
        Logging("Access archive menu")

        Waits.WaitUntilPageIsLoaded(page_xpath=data["archive"]["archive_leftside"])
        Commands.ReloadBrowser(data["archive"]["archive_leftside"])
    
    try:
        Commands.FindElement(data["archive"]["active_submenu_li"].replace("[folder]", folder))
    except WebDriverException:
        Commands.ClickElement(data["archive"]["left_arrow_submenu"].replace("[folder]", folder))
    finally:
        Logging("Acces archive sub menu")
    
    if folder == "Company Archive":
        folder_xpath = data["archive"]["companyarchive_folder"].replace("[folder]", folder).replace("[folder_name]", folder_name)
    else:
        folder_xpath = data["archive"]["myarchive_folder"].replace("[folder]", folder).replace("[folder_name]", folder_name)
    
    Commands.Wait10s_ClickElement(folder_xpath)
    Logging("Access my archive folder")

    try:
        Waits.WaitElementLoaded(5, data["archive"]["secure_input"])
        Commands.InputElement_2Values(data["archive"]["secure_input"], data["security_password"], Keys.ENTER)
        Logging("Input security password")
        Waits.Wait10s_ElementLoaded("//*[@id='set-my-arch']")
    except WebDriverException:
        pass  

    Waits.Wait10s_ElementLoaded(data["archive"]["list_item"])

    sort_xpath = data["archive"]["list_sorting"]
    sorting = Commands.FindElement(sort_xpath)
    sorting_value = Functions.GetElementAttribute(sort_xpath, "class")
    i=0
    for i in range(1,3):
        i+=1
        if "fa-sort-asc" in sorting_value:
            Waits.WaitUntilPageIsLoaded(None)
            break
        else:
            sorting.click()
    
    Waits.WaitUntilPageIsLoaded(None)
    doc_name = Functions.GetElementText(data["archive"]["list_item"])
    Commands.ClickElement(data["archive"]["list_item"])
    Logging("Copy to Archive: " + doc_name)
    
    if doc_name == document_name:
        Logging("Document is copied to archive successfully " + objects.testcase_pass)
    else:
        Logging("Fail to copy approval to archive " + objects.testcase_fail)

    archived_date = Functions.GetElementText(data["archive"]["archived_date"])
    Logging("archived_date.text " + archived_date)
    if archived_date != "-":
        Logging("Archived date is displayed normally " + objects.testcase_pass)
        today_date = objects.year + "/" + objects.month + "/" + objects.day
        if today_date == archived_date:
            Logging("Document is archived successfully " + objects.testcase_pass)
        else:
            Logging("Fail to archive document " + objects.testcase_fail)
    else:
        Logging("Archived date is not displayed " + objects.testcase_fail)

    Logging("View Archive - Open content")
    Commands.SwitchToFrame("//*[@id='viewDetail']")

    time.sleep(2)

    if "Title" in driver.page_source:
        Logging("View content of archived approval successfully " + objects.testcase_pass)
        result = True
    else:
        Logging("Fail to view content of archived approval " + objects.testcase_fail)
        result = False
        
    Commands.SwitchToDefaultContent()
    
    Commands.Wait10s_ClickElement(data["archive"]["view_close"])
    Logging("View Archive - Close view mode")

    Waits.Wait10s_ElementLoaded( data["archive"]["list_item"])

    return result

def CopyArchive_SelectArchiveFolder():
    Waits.Wait10s_ElementLoaded("//a[text()='My Archive']")
    archive_folder = []
    try:
        my_archive_folder = Commands.FindElement(data["archive"]["copy_archive_my"])
        archive_folder_name = my_archive_folder.text
        Logging("My archive folder name: " + archive_folder_name)
        my_archive_folder.click()
        Logging("-> Select my archive folder")
        archive_folder.append("My Archive")
        archive_folder.append(archive_folder_name)
    except WebDriverException:
        company_archive_xpath = data["archive"]["copy_archive_company"]
        Commands.ClickElement(company_archive_xpath)
        Logging("Open company archive")

        try:
            Waits.WaitElementLoaded(3, data["archive"]["copy_archive_company_folder"])
            company_archive_folder = Commands.FindElement(data["archive"]["copy_archive_company_folder"])
            archive_folder_name = company_archive_folder.text
            Logging("Company archive folder name: " + archive_folder_name)
            company_archive_folder.click()
            Logging("-> Select company archive folder")
            archive_folder.append("Company Archive")
            archive_folder.append(archive_folder_name)
        except WebDriverException:
            archive_folder_name = None
            Logging("Cannot find archive folder")
    
    time.sleep(1)

    if bool(archive_folder_name) == True:
        Commands.ClickElement(data["archive"]["save_copy_archive"])
        Logging("Click Save button")
        try:
            Waits.Wait10s_ElementLoaded("//div[contains(., 'Selected data')]")
            Logging("Copy to archive successfully " + objects.testcase_pass)
        except WebDriverException:
            Logging("Fail to copy archive " + objects.testcase_fail)
        finally:
            Commands.ClickElement(data["resource"]["close_warning"])
            Logging("Close warning")
    else:
        Commands.ClickElement(data["archive"]["close_copy_archive"])
        Logging("Click Close button")
        PrintYellow("Cannot find folder to be selected for copy")
    
    return archive_folder

def DefineListLength(xpath):
    try:
        Waits.Wait10s_ElementLoaded(xpath)
        list_length = Functions.GetListLength(xpath)
    except WebDriverException:
        list_length = 0
    
    return list_length

def CollectExcelLog(filename, json_file):
    wb = load_workbook(filename)
    current_sheet = wb.active
    last_row = int(len(list(current_sheet.rows)))

    menu_dict = dict(data['testcase_result'])

    not_executed_tc = []
    failed_tc = []
    testcase_excel = []
    testcase_list = []
    row_number = 0

    for menu_name in menu_dict.keys():
        menu = dict(data['testcase_result'][menu_name])
        for testcase_name in menu.keys():
            menu_item = data["testcase_result"]["" + menu_name + ""]["" + testcase_name + ""]["pass"]["menu"]
            testcase_item = data["testcase_result"]["" + menu_name + ""]["" + testcase_name + ""]["pass"]["testcase"]
            testcase = menu_item + " : " + testcase_item
            testcase_list.append(testcase)

    for row_number in range(1, last_row):
        row_number += 1
        
        menu_name_excel = current_sheet.cell(row=row_number, column=1).value
        testcase_name_excel = current_sheet.cell(row=row_number, column=3).value
        testcase_status_excel = current_sheet.cell(row=row_number, column=4).value
        
        testcase_excel_item = menu_name_excel + " : " + testcase_name_excel
        testcase_excel.append(testcase_excel_item)
        
        if "Fail" in testcase_status_excel:
            failed_tc.append(testcase_excel_item)

    for checked_testcase in testcase_list:
        if checked_testcase not in testcase_excel:
            not_executed_tc.append(checked_testcase)
    
    log = {
        "error_log": not_executed_tc,
        "fail_log": failed_tc
    }

    return log

def DefineTestCaseResult(testcase_name, status, error_list):
    if status == "complete":
        Logging("Test case [" + testcase_name + "] is executed completely ")
    else:
        Logging("An error occurs while executing Test case [" + testcase_name + "]")
        list(error_list).append(testcase_name)

def SearchDetailsBySelectBox(**search_details):
    listfooter_xpath = search_details["listfooter_xpath"]
    search_dict = search_details["search_dict"]
    expander_xpath = search_details["expander_xpath"]
    selectbox_xpath = search_details["selectbox_xpath"]
    searchinput_xpath = search_details["searchinput_xpath"]
    search_button_xpath = search_details["search_button_xpath"]
    menu_name = search_details["menu_name"]
    
    list1 = CollectListData(list_footer=listfooter_xpath, page_total="")
    
    result_dict = {}
    search_result = None
    if list1["total_items"] == 1:
        # Input random key to check search function
        # Key: fggfggfdgfd => Result should be No data
        # If list does not change => search function does not work
        
        search_data = "title"
        current_url = DefineCurrentURL()
        if "asset" in current_url:
            selectbox_value = dict(search_dict)["name"]["value"]
        else:
            selectbox_value = dict(search_dict)["title"]["value"]
        search_key = "fggfggfdgfd"

        Commands.ClickElement(expander_xpath)
        Logging("Click search expand button")
        
        Waits.Wait10s_ElementLoaded(selectbox_xpath)
        time.sleep(1)
        
        Commands.Selectbox_ByValue(selectbox_xpath, selectbox_value)
        Logging("Select search field: [" + selectbox_value + "]")

        Waits.Wait10s_ElementLoaded(searchinput_xpath)
        time.sleep(1)
        
        Commands.InputElement(searchinput_xpath, search_key)
        Logging("Send key for searching: [" + search_key + "]")

        Commands.ClickElement(search_button_xpath)
        Logging("Click Search button")

        try:
            Waits.Wait10s_ElementLoaded(data["common"]["list_nodata"])
            Logging("Search successfully [List is No Data]")
            search_result = True
        except WebDriverException:
            search_result = False
    elif list1["total_items"] > 1:
        for search_data in dict(search_dict).keys():
            selectbox_value = dict(search_dict)[search_data]["value"]
            search_key = dict(search_dict)[search_data]["key"]

            Commands.ClickElement(expander_xpath)
            Logging("Click search expand button")
            
            Waits.Wait10s_ElementLoaded(selectbox_xpath)
            time.sleep(1)
            
            Commands.Selectbox_ByValue(selectbox_xpath, selectbox_value)
            Logging("Select search field: [" + selectbox_value + "]")

            Waits.Wait10s_ElementLoaded(searchinput_xpath)
            time.sleep(1)
            
            Commands.InputElement(searchinput_xpath, search_key)
            Logging("Send key for searching: [" + search_key + "]")

            Commands.ClickElement(search_button_xpath)
            Logging("Click Search button")

            if list1["total_items"] > 1:
                i=0
                for i in range(1,10):
                    i+=1
                    time.sleep(1)
                    list2 = CollectListData(list_footer=listfooter_xpath, page_total="")
                    #Logging("list2 " + str(list2["total_items"]))
                    if list2["total_items"] != list1["total_items"]:
                        search_result = True
                        Logging("Search successfully")
                        break
                    else:
                        search_result = False
            else:
                try:
                    if "No Data" in driver.page_source:
                        Logging("List is empty")
                        search_result = False
                except:
                    pass
    else:
        search_result = False
        
    try:
        Commands.FindElement("//div[@class='error-container-wrap']")
        Logging("Page is error")
        TestCase_LogResult(**data["testcase_result"]["" + menu_name + ""]["search"]["fail"])
    except WebDriverException:
        if search_result == True:
            result_dict[search_data] = True
        else:
            result_dict[search_data] = False
            Logging("Cannot define search result")
        list3 = CollectListData(list_footer=listfooter_xpath, page_total="")
        #Logging("list3 " + str(list3["total_items"]))

        search_input_xpath = expander_xpath + "/preceding-sibling::input"
        Commands.InputElement(search_input_xpath, Keys.RETURN)
        Logging("Clear key in search box")
        i=0
        for i in range(0,10):
            i+=1
            time.sleep(1)
            list4 = CollectListData(list_footer=listfooter_xpath, page_total="")
            #Logging("list4 " + str(list4["total_items"]))
            if list4["total_items"] != list3["total_items"]:
                reset = True
                break
            else:
                reset = False
        if reset == True:
            Logging("Reset search result successfully")
        else:
            Logging("Fail to reset list")

        expander = Commands.FindElement(expander_xpath)
        expender_attr = Functions.GetElementAttribute(expander_xpath, "aria-expanded")
        if "true" in expender_attr:
            expander.click()
            Logging("Close search expander")
            
        time.sleep(1)
    
    for search_label in result_dict.keys():
        if result_dict[search_label] == False:
            dict(data["testcase_result"]["" + menu_name + ""]["search"]["fail"]).update({"description": "Fail to search with " + search_label})
            TestCase_LogResult(**data["testcase_result"]["" + menu_name + ""]["search"]["fail"])
    
    if False not in result_dict.values():
        TestCase_LogResult(**data["testcase_result"]["" + menu_name + ""]["search"]["pass"])

def ValidateUnexpectedModal():
    try:
        Commands.ClickElement("//div[@class='modal-dialog ui-draggable']/div/div[contains(@class, 'modal-header')]/button")
        Logging("Close alert modal")
        Waits.Wait10s_ElementInvisibility("//div[@class='modal-dialog ui-draggable']/div/div[contains(@class, 'modal-header')]/button")
        alert = True
    except WebDriverException:
        alert = False
    
    if alert == True:
        current_url = DefineCurrentURL()
        menu = current_url.split("/#/")[1].split("/")[0]
        if "addrbook" in current_url:
            menu = "contact"
        write_back_xpath = data[menu]["write_back_xpath"]
        view_back_xpath = data[menu]["view_back_xpath"]
        try:
            view_path = menu + data[menu]["view_back_xpath"]
            if view_path in current_url:
                Commands.ClickElement(view_back_xpath)
                Logging("Back from view page")
            else:
                Commands.ClickElement(write_back_xpath)
                Logging("Back from write page")
        except WebDriverException:
            pass
        finally:
            Waits.WaitUntilPageIsLoaded(page_xpath=None)