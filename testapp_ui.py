from distutils.log import error
import time, os
import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from common_functions import *
from groupware_execution import RunMainFeatures


def Messages(status):
    if status == "pass":
        msg = "Execution is executed successfully."
    else:
        msg = "There is an error while executing. Please run again or report your issue."

    messagebox.showinfo(title='Information', message=msg)    

def RemoveTextDomain(event):
    domain_value_text = domain_text.get()
    if domain_value_text == tooltips_dict["domain"]:
        domain_value.delete(0, END)
        domain_value.config(foreground="#000000")

def RemoveTextId(event):
    id_value_text = id_text.get()
    if id_value_text == tooltips_dict["id"]:
        id_value.delete(0, END)
        id_value.config(foreground="#000000")

def RemoveTextPw(event):
    pw_value_text = pw_text.get()
    if pw_value_text == tooltips_dict["pw"]:
        pw_value.delete(0, END)
        pw_value.config(foreground="#000000")

def RemoveTextEmail(event):
    email_value_text = email_text.get()
    if email_value_text == tooltips_dict["email"]:
        email_value.delete(0, END)
        email_value.config(foreground="#000000")

def RemoveTextRecipient(event):
    recipient_value_text = recipient_text.get()
    if recipient_value_text == tooltips_dict["recipient"]:
        recipient_value.delete(0, END)
        recipient_value.config(foreground="#000000")

def ShowTextDomain(event):
    domain_value.config(state=NORMAL)
    domain_value_text = domain_text.get()

    if bool(domain_value_text) == False:
        domain_value.insert(0, "Domain name")
        domain_value.config(foreground="#d3d3d3")

def ShowTextId(event):
    id_value.config(state=NORMAL)
    id_value_text = id_text.get()

    if bool(id_value_text) == False:
        id_value.insert(0, "User id for login")
        id_value.config(foreground="#d3d3d3")

def ShowTextPw(event):
    pw_value.config(state=NORMAL)
    pw_value_text = pw_text.get()

    if bool(pw_value_text) == False:
        pw_value.insert(0, "User pw for login")
        pw_value.config(foreground="#d3d3d3")

def ShowTextEmail(event):
    email_value.config(state=NORMAL)
    email_value_text = email_text.get()

    if bool(email_value_text) == False:
        email_value.insert(0, "Email address to receive test result")
        email_value.config(foreground="#d3d3d3")

def ShowTextRecipient(event):
    recipient_value.config(state=NORMAL)
    recipient_value_text = recipient_text.get()

    if bool(recipient_value_text) == False:
        recipient_value.insert(0, "Recipient id/name for task report recipient / approver of approval menu")
        recipient_value.config(foreground="#d3d3d3")

def ShowTextImport(event):
    import_value.config(state=NORMAL)

    if bool(import_text.get()) == False:
        import_value.insert(0, data["tooltip"]["import"])
        import_value.config(foreground="#d3d3d3")

def RemoveTextImport(event):
    if import_text.get() == data["tooltip"]["import"]:
        import_value.delete(0, END)
        import_value.config(foreground="#000000")

def ShowTextPlan(event):
    plan_value.config(state=NORMAL)

    if bool(plan_text.get()) == False:
        plan_value.insert(0, data["tooltip"]["plan"])
        plan_value.config(foreground="#d3d3d3")

def RemoveTextPlan(event):
    if plan_text.get() == data["tooltip"]["plan"]:
        plan_value.delete(0, END)
        plan_value.config(foreground="#000000")

def CheckboxAll():
    global checkbox_list
    checkbox_list = [mail_checkbox, board_checkbox, contact_checkbox, calendar_checkbox, 
                    approval_checkbox, circular_checkbox, todo_checkbox, archive_checkbox, 
                    asset_checkbox, work_diary_checkbox, task_report_checkbox, project_checkbox, 
                    resource_checkbox, expense_checkbox, clouddisk_checkbox]
    
    for checkbox in checkbox_list:
        if all_folders.get() == True:
            checkbox.select()
        else:
            checkbox.deselect()

def CheckFolders():
    menu_dict = {
        "Mail": {
            "value": mail.get(),
            "button": mail_checkbox
        },
        "Board": {
            "value": board.get(),
            "button": board_checkbox
        },
        "Contact": {
            "value": contact.get(),
            "button": contact_checkbox
        },
        "Calendar": {
            "value": calendar.get(),
            "button": calendar_checkbox
        },
        "Approval": {
            "value": approval.get(),
            "button": approval_checkbox
        },
        "Circular": {
            "value": circular.get(),
            "button": circular_checkbox
        },
        "ToDo": {
            "value": todo.get(),
            "button": todo_checkbox
        },
        "Archive": {
            "value": archive.get(),
            "button": archive_checkbox
        },
        "Asset": {
            "value": asset.get(),
            "button": asset_checkbox
        },
        "Work Diary": {
            "value": work_diary.get(),
            "button": work_diary_checkbox
        },
        "Task Report": {
            "value": task_report.get(),
            "button": task_report_checkbox
        },
        "Project": {
            "value": project.get(),
            "button": project_checkbox
        },
        "Resource": {
            "value": resource.get(),
            "button": resource_checkbox
        },
        "Expense": {
            "value": expense.get(),
            "button": expense_checkbox
        },
        "CloudDisk": {
            "value": clouddisk.get(),
            "button": clouddisk_checkbox
        },
        "Whisper": {
            "value": whisper.get(),
            "button": whisper_checkbox
        }
    }

    for folder_name in menu_dict.keys():
        if menu_dict[folder_name]["value"] == False and all_folders.get() == True:
            all_checkbox.deselect()
    
    return menu_dict

def QuitExecution():
    root.destroy()
    try:
        driver.quit()
    except:
        pass

def Selection_Value():
    selected_value = str(test_type_var.get())

    return selected_value

def ValidateExecutionData(domain_submit, id_submit, pw_submit, email_submit, recipient_submit):
    assigned_test = Selection_Value()
    if assigned_test == "1":
        try:
            #RunBasicTest(domain_submit, id_submit, pw_submit, email_submit)
            execution_status = True
        except:
            execution_status = False

        #RunBasicTest(domain_submit, id_submit, pw_submit, email_submit)
        #execution_status = True
    else:
        if recipient_submit == tooltips_dict["recipient"]:
            recipient_submit = None

        try:
            returned_msg = RunMainFeatures(domain_submit, id_submit, pw_submit, email_submit, recipient_submit)
            if returned_msg == False:
                messagebox.showinfo(title='Information', message="The automation test for main features is currently not supported. You can switch to basic test")
            execution_status = True
        except:
            execution_status = False
        
        #returned_msg = RunMainFeatures(domain_submit, id_submit, pw_submit, email_submit, recipient_submit)
        #execution_status = True

    return execution_status

def CloseImportModal():
    import_frame.destroy()

def ReadExcelFile(file_path):

    wb = load_workbook(file_path)
    current_sheet = wb.active
    last_row = int(len(list(current_sheet.rows)))

    row_number=0

    for row_number in range(1, last_row):
        row_number+=1
        
        domain = current_sheet.cell(row=row_number, column=1).value
        if bool(domain) == True:
            login_dict["domain"].append(domain)
        else:
            login_dict["domain"].append(None)
        
        id = current_sheet.cell(row=row_number, column=2).value
        if bool(id) == True:
            login_dict["id"].append(id)
        else:
            login_dict["id"].append(None)

        pw = current_sheet.cell(row=row_number, column=3).value
        if bool(pw) == True:
            login_dict["pw"].append(pw)
        else:
            login_dict["pw"].append(None)

        recipient = current_sheet.cell(row=row_number, column=4).value
        if bool(recipient) == True:
            login_dict["recipient"].append(recipient)
        else:
            login_dict["recipient"].append(None)

    return login_dict

def SelectFile():
    global file_path, file_path_text
    file_path = filedialog.askopenfile(initialdir="/Downloads", filetypes=(("excel files", ".xls"), ("excel files", ".xlsx")))
    file_path_text = str(file_path).replace("<_io.TextIOWrapper name='", "").replace("' mode='r' encoding='cp1252'>", "")
    
    #selected_file_label.config(text=file_path_text)

    if bool(file_path_text) == True:
        allowed_files = ["xls", "xlsx"]
        selected_file_extension = file_path_text.split(".")[-1]
        if selected_file_extension not in allowed_files:
            messagebox.showinfo(title="Info", message="file extension is not accepted")
        else:
            login_dict = ReadExcelFile(file_path_text)
            length = int(len(list(login_dict["domain"])))
            
            import_domain_value = ""
            import_id_value = ""
            import_pw_value = ""
            import_recipient_value = ""

            i = 0
            for i in range(0, length):
                output_domain = login_dict["domain"][i]
                output_id = login_dict["id"][i]
                output_pw = login_dict["pw"][i]
                output_recipient = login_dict["recipient"][i]

                if i == 0:
                    import_domain_value = import_domain_value + output_domain
                    import_id_value = import_id_value + output_id
                    import_pw_value = import_pw_value + output_pw
                    import_recipient_value = import_recipient_value + output_recipient
                else:
                    import_domain_value = import_domain_value + "," + output_domain
                    import_id_value = import_id_value + "," + output_id
                    import_pw_value = import_pw_value + "," + output_pw
                    import_recipient_value = import_recipient_value + "," + output_recipient
                
                #data_table.insert(parent='',index='end',iid=i,text='', values=(output_domain, output_id, output_pw, output_recipient))

                i+=1
            
            domain_value.delete(0, END)
            domain_value.insert(0, import_domain_value)
            domain_value.config(foreground="#000000", state="disabled")
            
            id_value.delete(0, END)
            id_value.insert(0, import_id_value)
            id_value.config(foreground="#000000", state="disabled")

            pw_value.delete(0, END)
            pw_value.insert(0, import_pw_value)
            pw_value.config(foreground="#000000", state="disabled")

            recipient_value.delete(0, END)
            recipient_value.insert(0, import_recipient_value)
            recipient_value.config(foreground="#000000", state="disabled")

def ValidateExecutionFields(domain_submit, id_submit, pw_submit, email_submit, recipient_submit):
    tooltips = [
        "Domain name",
        "User id for login",
        "User pw for login",
        "Email address to receive test result",
        "Recipient id/name as task report recipient / approver of approval menu"
    ]

    required_fields = [domain_submit, id_submit, pw_submit, email_submit]
    enable_run = []

    for required_field in required_fields:
        if required_field in tooltips:
            # => If field input value == tooltip text
            # => field value is empty (user did not fill in)
            enable_run.append(False)
            warning_msg = "Required fields (*) cannot be empty"
        else:
            if "," in domain_submit:
                import_duplicate = []
                for import_data in required_fields:
                    import_duplicate.append(len(import_data))
                import_duplicate.append(len(recipient_submit))
                import_duplicate = list(dict.fromkeys(import_duplicate))
                
                if len(import_duplicate) < 4:
                    enable_run.append(True)
                else:
                    enable_run.append(False)
                    warning_msg = "Please check the import data"
            else:
                enable_run.append(True)

    if recipient_submit in tooltips:
        recipient_submit = None
    
    if False not in enable_run:
        execution_status = ValidateExecutionData(domain_submit, id_submit, pw_submit, email_submit, recipient_submit)
    else:
        messagebox.showinfo("Warning", warning_msg)
        execution_status = None

    return execution_status

def RunImportExecution():
    driver = Driver.DefineDriver("web")

    # Get value user input for test data
    domain_submit = domain_text.get()
    id_submit = id_text.get()
    pw_submit = pw_text.get()
    email_submit = email_text.get()
    recipient_submit = recipient_text.get()

    execution_status_list = []
    if "," in domain_submit:
        # User run multiple test data splitted by ","
        domain_submit_list = list(domain_submit.split(","))
        id_submit_list = list(id_submit.split(","))
        pw_submit_list = list(pw_submit.split(","))
        recipient_submit_list = list(recipient_submit.split(","))
        for list_ele in domain_submit_list:
            i = domain_submit_list.index(list_ele)
            domain_submit = domain_submit_list[i]
            id_submit = id_submit_list[i]
            pw_submit = pw_submit_list[i]
            recipient_submit = recipient_submit_list[i]
            execution_status = ValidateExecutionFields(domain_submit, id_submit, pw_submit, email_submit, recipient_submit)
            if execution_status != None:
                execution_status_list.append(execution_status)
            open(Files.execution_log, "w")
            time.sleep(1)
    else:
        execution_status = ValidateExecutionFields(domain_submit, id_submit, pw_submit, email_submit, recipient_submit)
        if execution_status != None:
            execution_status_list.append(execution_status)
    
    if False not in execution_status_list:
        Messages("pass")
        driver.quit()
    else:
        Messages("fail")

    Files.ConfigFiles()

def ImportFunction():
    global import_frame
    import_frame = tk.Tk()
    screen_width = import_frame.winfo_screenwidth()
    screen_height = import_frame.winfo_screenheight()
    width = 610
    height = 400
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    import_frame.geometry('%dx%d+%d+%d' % (width, height, x, y))
    import_frame.resizable(True, True)
    import_frame.title('Import modal')
    import_frame.config(padx=10, pady=10)

    import_table = ttk.Frame(import_frame)
    import_table.pack(padx=10, pady=10, fill='x', expand=True)
    
    import_button = ttk.Button(import_table, text="Select File", width=25, command=SelectFile)
    import_button.grid(row=1, column=1, padx=15)

    login_button = ttk.Button(import_table, text="LogIn", width=25, command=RunImportExecution)
    login_button.grid(row=1, column=2, padx=15)

    close_button = ttk.Button(import_table, text="Close", width=25, command=CloseImportModal)
    close_button.grid(row=1, column=3, padx=15)
    
    global file_path_text
    if file_path_text:
        file_path_text = os.path.abspath(file_path.name)

    global selected_file_label
    selected_file_label = ttk.Label(import_frame, text=file_path_text)
    selected_file_label.pack()

    global email_value
    email_text = tk.StringVar()

    email_label = ttk.Label(import_frame, text="(*) Receive Email")
    email_label.pack(fill='x', expand=True)

    email_value = ttk.Entry(import_frame, textvariable=email_text)
    email_value.pack(fill='x', expand=True)

    global data_table
    data_table = ttk.Treeview(import_frame, selectmode ='browse')
    data_table.pack()

    data_table['columns'] = ("domain", "id", "pw", "recipient")
    data_table['show'] = 'headings'

    data_table.column("domain",anchor=CENTER, width=150)
    data_table.column("id",anchor=CENTER, width=150)
    data_table.column("pw",anchor=CENTER, width=150)
    data_table.column("recipient",anchor=CENTER, width=150)

    data_table.heading("domain",anchor=CENTER, text="Domain")
    data_table.heading("id",anchor=CENTER, text="ID")
    data_table.heading("pw",anchor=CENTER, text="PW")
    data_table.heading("recipient",anchor=CENTER, text="Recipient")

    #import_frame.mainloop()

def StartFunction():
    # Reset excel log before checking new selected folder
    Logs.ClearExcel_CollectMenu()

    # Clear logs
    lb.delete(4,END)
    
    # Validated selected menu from user
    selected_menu = []
    menu_dict = CheckFolders()
    
    for menu in menu_dict.keys():
        if menu_dict[menu]["value"] == True:
            selected_menu.append(menu)

    # Update new selected menu list in excel log
    Logs.UpdateSelection_CollectMenu(selected_menu)

    # Start webdriver and run test
    user_input = {
        "domain_name": domain_text.get(),
        "user_id": id_text.get(),
        "user_pw": pw_text.get(),
        "date": objects.date_time
    }
    Logs.UserInput(**user_input)

    testplan_name = "Test App"
    build_name = objects.date_id
    
    domain_config = {
        "domain_name": "https://%s/ngw/app/#" % str(domain_text.get()),
        "user_id": id_text.get(),
        "user_pw": pw_text.get()
    }

    if recipient_text.get() == tooltips_dict["recipient"]:
        domain_config["recipient_id"] = False
        domain_config["recipient_name"] = False
    else:
        domain_config["recipient_id"] = recipient_text.get()
        domain_config["recipient_name"] = recipient_text.get()

    RunMainFeatures(**domain_config)

    # Define which menu failed while running
    error_list = Logs.CheckResult_CollectMenu()
    if bool(error_list) == True:
        messagebox.showerror("Error", "Menu list failed while running %s" % str(error_list))
    else:
        messagebox.showinfo("Success", "Execution finished")
    
    InsertLogs()

def SubmitTestPlan():
    ''' When user submit the test plan, write the test plan name in txt file
        validation: if user input the same value -> show error
                    else update the test plan name in txt file
                    -> value of test plan name in txt file to be used as file name for test case log result '''

    testplan_name = plan_text.get()
    current_testplan_name = open(Files.testplan, "r").read()
    if testplan_name != current_testplan_name:
        testplan_file = open(Files.testplan, "w")
        testplan_file.write(testplan_name)
        testplan_file.close()
        print("# update test plan file")
        messagebox.showinfo("Success", "Create plan successfully! You can view your test log on web from now")
    else:
        messagebox.showwarning("Warning", "This test plan already exists!")

file_path_text = None
file_path = None
log_msg = None

tooltips_dict = data["tooltip"]

login_dict = {
    "domain": [],
    "id": [],
    "pw": [],
    "recipient": []
    }

def show_frame(frame):
    frame.tkraise()

def SignInPage():
    global domain_text, id_text, pw_text, email_text, recipient_text, test_type_var, import_text, plan_text
    global all_folders, mail, board, contact, calendar, approval, circular, todo
    global archive, asset, work_diary, task_report, project, resource, expense, clouddisk, whisper

    domain_text = tk.StringVar()
    id_text = tk.StringVar()
    pw_text = tk.StringVar()
    email_text = tk.StringVar()
    recipient_text = tk.StringVar()
    test_type_var = IntVar()
    import_text = tk.StringVar()
    plan_text = tk.StringVar()

    all_folders = tk.BooleanVar()
    mail = tk.BooleanVar()
    board = tk.BooleanVar()
    contact = tk.BooleanVar()
    calendar = tk.BooleanVar()
    approval = tk.BooleanVar()
    circular = tk.BooleanVar()
    todo = tk.BooleanVar()
    archive = tk.BooleanVar()
    asset = tk.BooleanVar()
    work_diary = tk.BooleanVar()
    task_report = tk.BooleanVar()
    project = tk.BooleanVar()
    resource = tk.BooleanVar()
    expense = tk.BooleanVar()
    clouddisk = tk.BooleanVar()
    whisper = tk.BooleanVar()

    # Sign in frame
    signin = Frame(root, width=600)
    #signin = Frame(self)
    signin.pack(padx=10, pady=10, side=LEFT, expand=True)

    global domain_value, id_value, pw_value, recipient_value, email_value, recipient_value, import_value, plan_value

    #
    domain_label = ttk.Label(signin, text="(*) Domain")
    domain_label.grid(column=0, row=1, pady=2, sticky="W")

    placeholder_domain = tooltips_dict["domain"]
    domain_value = ttk.Entry(signin, textvariable=domain_text)
    domain_value.insert(0, placeholder_domain)
    domain_value.config(foreground="#d3d3d3")
    domain_value.bind("<FocusIn>", RemoveTextDomain)
    domain_value.bind("<FocusOut>", ShowTextDomain)
    domain_value.grid(column=1, row=1, ipadx=180, ipady=3, pady=7, columnspan=7, sticky="W")

    #
    id_label = ttk.Label(signin, text="(*) User ID")
    id_label.grid(column=0, row=2, pady=2, sticky="W")

    placeholder_id = tooltips_dict["id"]
    id_value = ttk.Entry(signin, textvariable=id_text)
    id_value.insert(0, placeholder_id)
    id_value.config(foreground="#d3d3d3")
    id_value.bind("<FocusIn>", RemoveTextId)
    id_value.bind("<FocusOut>", ShowTextId)
    id_value.grid(column=1, row=2, ipadx=180, ipady=3, pady=7, columnspan=7, sticky="W")

    #
    pw_label = ttk.Label(signin, text="(*) User PW")
    pw_label.grid(column=0, row=3, pady=2, sticky="W")

    placeholder_pw = tooltips_dict["pw"]
    pw_value = ttk.Entry(signin, textvariable=pw_text)
    pw_value.insert(0, placeholder_pw)
    pw_value.config(foreground="#d3d3d3")
    pw_value.bind("<FocusIn>", RemoveTextPw)
    pw_value.bind("<FocusOut>", ShowTextPw)
    pw_value.grid(column=1, row=3, ipadx=180, ipady=3, pady=7, columnspan=7, sticky="W")

    #
    recipient_label = ttk.Label(signin, text="Recipient")
    recipient_label.grid(column=0, row=5, pady=2, sticky="W")

    placeholder_recipient = tooltips_dict["recipient"]
    recipient_value = ttk.Entry(signin, textvariable=recipient_text)
    recipient_value.insert(0, placeholder_recipient)
    recipient_value.config(foreground="#d3d3d3")
    recipient_value.bind("<FocusIn>", RemoveTextRecipient)
    recipient_value.bind("<FocusOut>", ShowTextRecipient)
    recipient_value.grid(column=1, row=5, ipadx=180, ipady=3, pady=7, columnspan=7, sticky="W")

    # types of test
    test_type = ttk.Label(signin, text="Type of Test")
    test_type.grid(column=0, row=6, pady=2, sticky="W")

    basic_option_text = "Basic test"
    main_option_text = "Advanced test"

    basic_option = ttk.Radiobutton(signin, text=basic_option_text, variable=test_type_var, value=1, command=Selection_Value)
    basic_option.grid(column=1, row=6, columnspan=3, pady=7, sticky="W")

    main_option = ttk.Radiobutton(signin, text=main_option_text, variable=test_type_var, value=2, command=Selection_Value)
    main_option.grid(column=4, row=6, columnspan=4, pady=7, sticky="W")

    test_type_var.set(2)

    folders_label = ttk.Label(signin, text="Menu")
    folders_label.grid(column=0, row=7, sticky="W")

    global          all_checkbox, mail_checkbox, board_checkbox, contact_checkbox, calendar_checkbox
    global          approval_checkbox, circular_checkbox, todo_checkbox, archive_checkbox
    global          asset_checkbox, work_diary_checkbox, task_report_checkbox, project_checkbox
    global          resource_checkbox, expense_checkbox, clouddisk_checkbox, whisper_checkbox

    all_checkbox = tk.Checkbutton(signin, text="All", var=all_folders, command=CheckboxAll)
    all_checkbox.grid(column=1, row=7, sticky="W")

    mail_checkbox = tk.Checkbutton(signin, text="Mail", var=mail, command=CheckFolders)
    mail_checkbox.grid(column=2, row=7, sticky="W")

    board_checkbox = tk.Checkbutton(signin, text="Board", var=board, command=CheckFolders)
    board_checkbox.grid(column=3, row=7, sticky="W")

    contact_checkbox = tk.Checkbutton(signin, text="Contact", var=contact, command=CheckFolders)
    contact_checkbox.grid(column=4, row=7, sticky="W")

    calendar_checkbox = tk.Checkbutton(signin, text="Calendar", var=calendar, command=CheckFolders)
    calendar_checkbox.grid(column=5, row=7, sticky="W")

    approval_checkbox = tk.Checkbutton(signin, text="Approval", var=approval, command=CheckFolders)
    approval_checkbox.grid(column=1, row=8, sticky="W")

    circular_checkbox = tk.Checkbutton(signin, text="Circular", var=circular, command=CheckFolders)
    circular_checkbox.grid(column=2, row=8, sticky="W")

    todo_checkbox = tk.Checkbutton(signin, text="To-Do", var=todo, command=CheckFolders)
    todo_checkbox.grid(column=3, row=8, sticky="W")

    archive_checkbox = tk.Checkbutton(signin, text="Archive", var=archive, command=CheckFolders)
    archive_checkbox.grid(column=4, row=8, sticky="W")

    asset_checkbox = tk.Checkbutton(signin, text="Asset", var=asset, command=CheckFolders)
    asset_checkbox.grid(column=5, row=8, sticky="W")

    whisper_checkbox = tk.Checkbutton(signin, text="Whisper", var=whisper, command=CheckFolders)
    whisper_checkbox.grid(column=6, row=8, sticky="W")

    work_diary_checkbox = tk.Checkbutton(signin, text="Work Diary", var=work_diary, command=CheckFolders)
    work_diary_checkbox.grid(column=1, row=9, sticky="W")

    task_report_checkbox = tk.Checkbutton(signin, text="Task Report", var=task_report, command=CheckFolders)
    task_report_checkbox.grid(column=2, row=9, sticky="W")

    project_checkbox = tk.Checkbutton(signin, text="Project", var=project, command=CheckFolders)
    project_checkbox.grid(column=3, row=9, sticky="W")

    resource_checkbox = tk.Checkbutton(signin, text="Resource", var=resource, command=CheckFolders)
    resource_checkbox.grid(column=4, row=9, sticky="W")

    expense_checkbox = tk.Checkbutton(signin, text="Expense", var=expense, command=CheckFolders)
    expense_checkbox.grid(column=5, row=9, sticky="W")

    clouddisk_checkbox = tk.Checkbutton(signin, text="CloudDisk", var=clouddisk, command=CheckFolders)
    clouddisk_checkbox.grid(column=6, row=9, sticky="W")

    '''all_checkbox.select()
    CheckboxAll()'''

    # import button
    import_button = ttk.Label(signin, text="Import")
    import_button.grid(column=0, row=10, pady=2, sticky="W")

    placeholder_import = tooltips_dict["import"]
    import_value = ttk.Entry(signin, textvariable=import_text)
    import_value.insert(0, placeholder_import)
    import_value.config(foreground="#d3d3d3")
    import_value.bind("<FocusIn>", RemoveTextImport)
    import_value.bind("<FocusOut>", ShowTextImport)
    import_value.grid(column=1, row=10, ipadx=180, ipady=3, pady=7, columnspan=6, sticky="W")

    upload_button = ttk.Button(signin, text="Browse", command=SelectFile)
    upload_button.grid(column=6, row=10, ipadx=10, sticky="W")

    #
    plan_label = ttk.Label(signin, text="Create plan")
    plan_label.grid(column=0, row=11, pady=2, sticky="W")

    placeholder_plan = tooltips_dict["plan"]
    plan_value = ttk.Entry(signin, textvariable=plan_text)
    plan_value.insert(0, placeholder_plan)
    plan_value.config(foreground="#d3d3d3")
    plan_value.bind("<FocusIn>", RemoveTextPlan)
    plan_value.bind("<FocusOut>", ShowTextPlan)
    plan_value.grid(column=1, row=11, ipadx=180, ipady=3, pady=7, columnspan=6, sticky="W")

    submit_button = ttk.Button(signin, text="Submit", command=SubmitTestPlan)
    submit_button.grid(column=6, row=11, ipadx=10, sticky="W")

    # login button
    login_button = ttk.Button(signin, text="Start", width=20, command=StartFunction)
    login_button.grid(column=1, row=12, columnspan=8, padx=150, pady=20, ipady=4, sticky="W")

    return signin

def LogPage():
    canvas = Canvas(root, width=600, height=420)
    canvas.pack(padx=0, pady=10, side=RIGHT, expand=True)

    frame2 = Frame(canvas,relief=RIDGE)
    canvas.create_window(100, 10, anchor="nw", window=frame2, height=360)
    Name = tk.Label(frame2,text = "Project Name")
    Name.place(x=0,y=0)

    global lb
    lb = Listbox(frame2,font= (10),width=55,height=15)
    lb.pack(expand=True, fill=BOTH, side=LEFT)

    sb = Scrollbar(frame2, orient=VERTICAL)
    sb.pack(fill=Y, side=RIGHT)

    lb.configure(yscrollcommand=sb.set)
    sb.config(command=lb.yview)

    lb.insert(0, 'Project Name :')
    lb.insert(1,'Domain :')
    lb.insert(2,'Account :')
    lb.insert(3,'Version :')

    Save_logs = tk.Button(canvas,text = "Save logs",width=15)
    canvas.create_window(200, 380, anchor="nw", window=Save_logs)
    
    View_Web = tk.Button(canvas,text = "View on Web",width=15)
    canvas.create_window(380, 380, anchor="nw", window=View_Web)

    return canvas

def InsertLogs():
    wb = openpyxl.load_workbook(Files.testcase_log)
    current_sheet = wb.active

    row_length = len(list(current_sheet.rows))
    for row in range(1,row_length):
        row+=1 # start at 2
        row_insert = row + 2
        menu = current_sheet.cell(row=row, column=2).value
        testcase = current_sheet.cell(row=row, column=4).value
        status = current_sheet.cell(row=row, column=5).value

        if bool(menu) == True:
            if bool(status) == False:
                status = "Block"
            
            text_insert = "[%s] %s - %s" % (status, menu, testcase)
            lb.insert(row_insert, text_insert)
        
    wb.save(Files.testcase_log)

# root window
root = tk.Tk()
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
width = 1200
height = 450
x = (screen_width/3) - (width/2)
y = (screen_height/3) - (height/2)
root.geometry('%dx%d+%d+%d' % (width, height, x, y))
root.resizable(True, True)
root.title('Automation Test V1.5.1')

SignInPage()
LogPage()

root.mainloop()