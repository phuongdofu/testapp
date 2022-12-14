import socket, shutil, os, json, ast, time, platform, random
from openpyxl import load_workbook, Workbook
from flask_config import log_folder, testplan_id_file


# [Reference] https://www.digitalocean.com/community/tutorials/python-socket-programming-server-client
# [Reference] https://viblo.asia/p/lap-trinh-socket-bang-python-jvEla084Zkw

''' This file will be run at server '''

def ReceiveInput():
    # get the host name
    host = socket.gethostname()
    port = 5001  # initiate port no above 1024

    server_socket = socket.socket()  # get instance
    # look closely. The bind() function takes tuple as argument
    server_socket.bind((host, port))  # bind host address and port together

    # configure how many client the server can listen simultaneously
    server_socket.listen(2)
    conn, address = server_socket.accept()  # accept new connection
    print("Connection from: " + str(address))

    data = b''
    while True:
        data_chunk = conn.recv(1024)
        if data_chunk:
            data+=data_chunk
        else:
            break

    data = data.decode('utf-8')
    parsed_data = ast.literal_eval(str(data).replace("'", '"'))
    testcase_dict = dict(parsed_data)
    testplan_name = str(testcase_dict["1"]["tester"])
    section_id = str(testcase_dict["1"]["section_id"])
    testcase_filename = "%s_result_%s" % (testplan_name, section_id)

    current_path = os.path.dirname(os.path.realpath(__file__))
    if "\\" in current_path:
        file_name = current_path + '\\Log\\Test Log\\%s.xlsx' % testcase_filename
    else:
        file_name = current_path + '/Log/Test Log/%s.xlsx' % testcase_filename

    for file in os.listdir(log_folder):
        if not file.startswith(testplan_name): # Case name: Test Plan is not existing
            wb = load_workbook(testplan_id_file)
            ws = wb.active

            existing_id = []

            last_row = ws.max_row
            new_row = None

            for row in range(1,last_row):
                row+=1
                testplan_name_value = ws.cell(row=row, column=1).value
                testplan_id_value = ws.cell(row=row, column=2).value
                if bool(testplan_name_value) == False:
                    new_row = row
                    break
                else:
                    existing_id.append(testplan_id_value)
            
            new_id = int()
            while new_id < 1000:
                new_id = int(random.randint(1, 1000))
                if new_id not in existing_id:
                    break
                
            ws.cell(row=new_row, column=1).value = testplan_name
            ws.cell(row=new_row, column=2).value = new_id

    create_wb = Workbook()
    create_wb.save(file_name)
    
    time.sleep(1)
    
    # Write the received data in new excel file
    wb = load_workbook(file_name)
    ws = wb.active

    last_row = 88
    for row_number in range(1, last_row):
        if row_number == 1:
            ws.cell(row=row_number, column=1).value = "No"
            ws.cell(row=row_number, column=2).value = "Menu"
            ws.cell(row=row_number, column=3).value = "Submenu"
            ws.cell(row=row_number, column=4).value = "Test Case"
            ws.cell(row=row_number, column=5).value = "Status"
            ws.cell(row=row_number, column=6).value = "Date"
            ws.cell(row=row_number, column=7).value = "Tester"
        else:
            tc_id = str(row_number)
            ws.cell(row=row_number, column=1).value = tc_id
            ws.cell(row=row_number, column=2).value = testcase_dict[tc_id]["menu"]
            ws.cell(row=row_number, column=3).value = testcase_dict[tc_id]["submenu"]
            ws.cell(row=row_number, column=4).value = testcase_dict[tc_id]["testcase"]
            ws.cell(row=row_number, column=5).value = testcase_dict[tc_id]["status"]
            ws.cell(row=row_number, column=6).value = testcase_dict[tc_id]["date"]
            ws.cell(row=row_number, column=7).value = testcase_dict[tc_id]["tester"]

        row_number+=1

    wb.save(file_name)


if __name__ == '__main__':
    while True:
        ReceiveInput()