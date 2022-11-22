import socket, json
from common_functions import Files
from openpyxl import load_workbook

# [Reference] https://www.digitalocean.com/community/tutorials/python-socket-programming-server-client
# [Reference] https://viblo.asia/p/lap-trinh-socket-bang-python-jvEla084Zkw

''' This file will run at client '''

def SendTestCaseFile():
    host = "qavn.hanbiro.net"
    #host = socket.gethostname()
    port = 5000 

    client_socket = socket.socket()  # instantiate
    client_socket.connect((host, port))  # connect to the server

    api_data = {}

    file = Files.testcase_log
    wb = load_workbook(file)
    ws = wb.active

    last_row = 88
    for row_number in range(1,last_row):

        row_number+=1
        no = ws.cell(row=row_number, column=1).value
        menu = ws.cell(row=row_number, column=2).value
        submenu = ws.cell(row=row_number, column=3).value
        testcase = ws.cell(row=row_number, column=4).value
        status = ws.cell(row=row_number, column=5).value
        date = ws.cell(row=row_number, column=6).value
        tester = ws.cell(row=row_number, column=7).value
        section_id = ws.cell(row=2, column=9).value

        api_data[str(no)] = {
            "no": no,
            "menu": menu,
            "submenu": submenu,
            "testcase": testcase,
            "status": status,
            "date": date,
            "tester": tester,
            "section_id": section_id
        }

    wb.save(file)
    
    data_send = str(api_data).encode()
    client_socket.send(data_send)
    client_socket.close()

if __name__ == '__main__':
    SendTestCaseFile()