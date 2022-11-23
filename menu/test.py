from openpyxl import load_workbook

wb = load_workbook("D:\\PhuongDofu\\groupware-auto-test-2\\Attachment\\testcase_log_result.xlsx")
ws = wb.active

last_row = ws.max_row

print(last_row)